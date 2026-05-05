"""
Build AI_Site_Builder_Business_Plan_v2.xlsx from v1 with the post-Promto-analysis
strategy applied. Updates only the sheets that changed; preserves existing
formatting and column widths from the source template.
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DOWNLOADS = Path.home() / "Downloads"
SRC = DOWNLOADS / "AI_Site_Builder_Business_Plan_v1.xlsx"
DST = DOWNLOADS / "AI_Site_Builder_Business_Plan_v2.xlsx"

shutil.copyfile(SRC, DST)
wb = openpyxl.load_workbook(DST)

THIN = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", size=11, bold=True)
HEAD = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", start_color="1F2937")
ACCENT_FILL = PatternFill("solid", start_color="EEF0FF")
GOOD_FILL = PatternFill("solid", start_color="DCFCE7")
WARN_FILL = PatternFill("solid", start_color="FEF3C7")
WRAP = Alignment(wrap_text=True, vertical="top")


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row + 5)


def write_table(ws, title, subtitle, headers, rows, widths=None):
    clear_sheet(ws)
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 32

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEAD
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    for r_i, row in enumerate(rows, start=5):
        for c_i, val in enumerate(row, start=1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            c.alignment = WRAP
            c.font = THIN

    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# -------------------------------------------------------------------- #
# Sheet 0 — Позиционирование                                            #
# -------------------------------------------------------------------- #
ws = wb["0. Позиционирование"]
clear_sheet(ws)

ws["A1"] = "Omnia.AI — 0. Позиционирование (v2, post-Promto)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)

rows = [
    ("ONE-LINER", "Omnia.AI — это Promto + всё, чего ему не хватает, чтобы выйти из MVP в production: backend, выделенный сервер, бэкапы, mix LLM, версионирование с откатом, 152-ФЗ."),
    ("КАТЕГОРИЯ", "End-to-end production-платформа на AI для российского рынка. Не AI-чат, не конструктор без AI, не агентство — новая категория."),
    ("", ""),
    ("─── ПОЗИЦИОНИРОВАНИЕ vs КОНКУРЕНТЫ ───", ""),
    ("Promto.ai", "Они дают AI-чат + лендинги/чат-боты. Мы добавляем выделенный VPS, backend под ключ (FastAPI/Postgres), визуальный rollback, mix LLM, 152-ФЗ. Тот же entry-tier, шире функционал."),
    ("Lovable / v0 / Bolt", "Они не работают для РФ. Мы — рублёвая оплата, российские серверы, поддержка на русском, документы по 152-ФЗ."),
    ("Tilda / Wix", "У них нет AI. У нас AI + всё что у них."),
    ("Студия / агентство", "У них недели и 100к+. У нас минуты и 990 ₽/мес."),
    ("", ""),
    ("─── TAGLINE-КАНДИДАТЫ ───", ""),
    ("• Промпт. Сайт. Откат.", "ГЛАВНЫЙ — закрывает уникальное преимущество (rollback)"),
    ("• AI-кодинг с откатом и российским хостингом", "secondary"),
    ("• Сайт за 5 минут вместо 2 месяцев", "speed"),
    ("• 990 ₽ — production-стек на AI", "price-anchor"),
    ("", ""),
    ("─── ANTI-FEAR MESSAGING (для landing) ───", ""),
    ("«Слишком сложно для не-разработчика»", "Free 5 дней без карты + 5 шаблонов + видео-онбординг 2 мин."),
    ("«Promto уже есть, зачем переходить»", "Visual rollback + полный backend + 152-ФЗ + mix LLM. Та же цена — больше под капотом."),
    ("«AI сломает сайт»", "Визуальная лента версий, откат в 1 клик. Promto такого не делает."),
    ("«Сервис закроется — пропадут сайты»", "Экспорт ZIP/Docker в любой момент + опциональный GitHub-синк."),
    ("«Дорого по сравнению с Promto»", "Lite 990 ₽ ≈ Promto Старт 690 ₽. У нас за +300 ₽ — mix LLM, кастом-домен, версионирование."),
    ("", ""),
    ("─── PROOF POINTS ───", ""),
    ("• Серверный пул через SafeCloud / CORTEL — production-инфраструктура с дня 0", ""),
    ("• Mix LLM (Claude/GPT/Yandex/GigaChat/DeepSeek/Qwen) — vendor lock-in protection + 152-ФЗ", ""),
    ("• Визуальное версионирование — единственный конкурент, который делает AI-кодинг безопасным", ""),
    ("• Юнит-экономика: gross margin 62%, LTV/CAC ~10x, точка безубыточности ~860 платящих", ""),
    ("• Один счёт для клиента вместо 5", ""),
    ("• Российское юрлицо, рублёвая оплата, РФ-хостинг", ""),
]

for r_i, (k, v) in enumerate(rows, start=3):
    ws.cell(row=r_i, column=1, value=k).font = BOLD if k.startswith("─") or k in (
        "ONE-LINER", "КАТЕГОРИЯ"
    ) else THIN
    ws.cell(row=r_i, column=2, value=v).alignment = WRAP
    ws.cell(row=r_i, column=2).font = THIN
    ws.row_dimensions[r_i].height = 32 if v and len(str(v)) > 50 else 18

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 90

# -------------------------------------------------------------------- #
# Sheet 4 — Конкуренты (расширенная матрица с Promto)                  #
# -------------------------------------------------------------------- #
ws = wb["4. Конкуренты"]
write_table(
    ws,
    "Omnia.AI — 4. Конкуренты (v2, добавлен Promto.ai как ключевой конкурент)",
    "Сравнительная матрица. Зелёный = у нас лучше; жёлтый = паритет; красный = у конкурента лучше.",
    ["Параметр", "Omnia.AI", "Promto.ai", "Lovable / v0 / Bolt", "Tilda / Wix", "Webflow", "Студия"],
    [
        ["AI-генерация фронта", "да (mix LLM)", "да (Claude only)", "да", "нет", "нет", "нет"],
        ["Mix LLM (выбор модели)", "★ ДА — 6 моделей", "нет", "нет", "н/п", "н/п", "н/п"],
        ["Backend под ключ (Postgres+API+JWT)", "★ ДА — конкретно", "размыто (PHP-Redis)", "частично (Supabase)", "нет", "базовый", "да"],
        ["Auto-deploy", "да (SafeCloud)", "да (общий пул)", "через Vercel", "свой", "свой", "ручной"],
        ["Выделенный VPS клиенту", "★ ДА (Pro+)", "shared", "нет", "shared", "shared", "иногда"],
        ["Кастомный домен", "★ от Lite (990 ₽)", "только Pro+", "нет", "да", "да", "да"],
        ["Российские серверы (152-ФЗ)", "★ ДА", "нет", "нет", "сертифицир.", "нет", "зависит"],
        ["Российские LLM (Yandex/GigaChat)", "★ ДА", "нет", "нет", "н/п", "н/п", "н/п"],
        ["Рублёвая оплата (ЮKassa/Tinkoff)", "да", "да", "нет", "да", "нет", "да"],
        ["Free-тариф без карты", "★ ДА — 5 дней", "да", "trial", "да", "да", "нет"],
        ["Минимальная цена", "990 ₽", "690 ₽", "$20+", "от 12к/год", "$14/мес", "100к+ ₽"],
        ["Время до live", "минуты", "2-5 минут", "час + ручной deploy", "часы", "часы", "недели"],
        ["★ Визуальное версионирование с rollback", "★ ДА (1 клик)", "нет", "нет", "вручную", "нет", "git вручную"],
        ["Бэкапы БД (daily snapshot)", "★ ДА", "не упомянуто", "нет", "базово", "базово", "зависит"],
        ["Мониторинг + SLA", "★ 99.5-99.95%", "не упомянуто", "нет", "98%", "99%", "договорной"],
        ["Чат-боты (TG/VK)", "★ Pro+", "★ да", "нет", "нет", "нет", "да"],
        ["Бизнес-автоматизации", "Pro+", "★ да", "нет", "Zapier", "Zapier", "да"],
        ["1С-интеграция", "★ Enterprise", "★ упоминается", "нет", "плагины", "нет", "да"],
        ["Часы инженеров со скидкой", "★ Pro+", "нет", "нет", "через сторонних", "через сторонних", "вся работа"],
        ["Self-export (ZIP/Docker)", "★ ДА", "не упомянуто", "иногда", "нет", "нет", "сорсы выдаются"],
        ["GitHub-синк (опционально)", "★ Pro+", "не упомянуто", "иногда", "нет", "нет", "да"],
    ],
    widths=[44, 22, 22, 22, 18, 18, 18],
)

# Highlight "★" rows where we win
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=7):
    if "★" in str(row[1].value or ""):
        for cell in row:
            cell.fill = GOOD_FILL

# Append USP block
last_row = ws.max_row
ws.cell(row=last_row + 2, column=1, value="УНИКАЛЬНОЕ ПРЕДЛОЖЕНИЕ vs Promto.ai (3 главных аргумента)").font = HEAD
ws.cell(row=last_row + 2, column=1).fill = HEAD_FILL
ws.cell(row=last_row + 4, column=1, value="1. ★ Визуальный rollback").font = BOLD
ws.cell(row=last_row + 4, column=2, value="После каждого промпта — снапшот с превью. Откат в 1 клик. Promto не имеет.").alignment = WRAP

ws.cell(row=last_row + 6, column=1, value="2. Mix LLM (vendor lock-in protection)").font = BOLD
ws.cell(row=last_row + 6, column=2, value="Promto = только Anthropic. Мы = 6 провайдеров включая YandexGPT/GigaChat для 152-ФЗ. Если Anthropic ограничит РФ — Promto встанет.").alignment = WRAP

ws.cell(row=last_row + 8, column=1, value="3. Полный backend под ключ").font = BOLD
ws.cell(row=last_row + 8, column=2, value="FastAPI + Postgres 16 + JWT + Alembic-миграции конкретно. Promto = «PHP до Redis» расплывчато. На production-задачах он встанет, мы — нет.").alignment = WRAP


# -------------------------------------------------------------------- #
# Sheet 7 — Юнит-экономика (новая модель с 5 тарифами)                 #
# -------------------------------------------------------------------- #
ws = wb["7. Юнит-экономика"]
clear_sheet(ws)
ws["A1"] = "Omnia.AI — 7. Юнит-экономика (v2, post-Promto: 5 тарифов)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A2"] = "Синие — input. Чёрные — формулы. Жёлтый фон — ключевые предположения. Зелёный — целевые ориентиры."
ws["A2"].font = Font(italic=True, color="6B7280")

# Block 1: Globals
ws["A4"] = "БЛОК 1. Глобальные параметры"
ws["A4"].font = HEAD
ws["A4"].fill = HEAD_FILL

globals_block = [
    ("Курс ₽/$", 100),
    ("Доля output в общей генерации (output_share)", 0.5),
    ("Наценка на токены (markup)", 2.8),
    ("Среднее потребление токенов клиентом, М/мес", 4),
    ("Срок жизни клиента, мес", 12),
    ("Фиксированный OPEX в месяц (FOT + маркетинг + офис), ₽", 3500000),
    ("CAC blended, ₽", 3000),
]
for i, (k, v) in enumerate(globals_block, start=5):
    ws.cell(row=i, column=1, value=k)
    c = ws.cell(row=i, column=2, value=v)
    c.font = Font(color="0000FF")  # blue input
    c.fill = WARN_FILL

# Block 2: Tier mix (mature, M15+)
start = 13
ws.cell(row=start, column=1, value="БЛОК 2. Тарифные пакеты — adoption mix (зрелое, M15+)").font = HEAD
ws.cell(row=start, column=1).fill = HEAD_FILL

tier_headers = [
    "Тариф",
    "Цена ₽/мес",
    "Доля платящих",
    "Кошелёк ₽",
    "VPS себест. ₽",
    "Себест. серв. + LLM, ₽",
    "Маржа на тарифе, %",
]
for i, h in enumerate(tier_headers, start=1):
    c = ws.cell(row=start + 1, column=i, value=h)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

tiers = [
    ("Free (5 дней без карты)", 0, 0.0, 500, 0, "lead-magnet", "—"),
    ("Lite", 990, 0.50, 1000, 0, 600, "39%"),
    ("Starter", 2990, 0.25, 2500, 600, 1500, "50%"),
    ("Pro", 7990, 0.18, 6000, 3000, 5500, "31%"),
    ("Enterprise", 19990, 0.07, 18000, 8000, 14000, "30%"),
]
for i, t in enumerate(tiers, start=start + 2):
    for c_i, val in enumerate(t, start=1):
        cell = ws.cell(row=i, column=c_i, value=val)
        cell.font = THIN
        if c_i == 2 and isinstance(val, (int, float)):
            cell.number_format = '#,##0 ₽'

# ARPU calc
arpu_row = start + 8
ws.cell(row=arpu_row, column=1, value="Средневзв. ARPU платящих, ₽/мес").font = BOLD
ws.cell(row=arpu_row, column=2, value="=B16*C16+B17*C17+B18*C18+B19*C19").font = BOLD
ws.cell(row=arpu_row, column=2).fill = ACCENT_FILL
ws.cell(row=arpu_row, column=2).number_format = '#,##0 ₽'

# Block 3: Funnel & CAC
start = 24
ws.cell(row=start, column=1, value="БЛОК 3. Воронка (зрелое, M15+)").font = HEAD
ws.cell(row=start, column=1).fill = HEAD_FILL

funnel = [
    ("Visits на лендинг (мес)", 50000),
    ("Conversion → Free signup", 0.08),
    ("Free signups (мес)", "=B25*B26"),
    ("Conversion Free → Paid", 0.12),
    ("Paid customers (мес, новых)", "=B27*B28"),
    ("Retention M6", 0.60),
    ("CAC blended, ₽", 3000),
]
for i, (k, v) in enumerate(funnel, start=start + 1):
    ws.cell(row=i, column=1, value=k)
    cell = ws.cell(row=i, column=2, value=v)
    if isinstance(v, str) and v.startswith("="):
        cell.font = Font(color="000000")  # formula
    else:
        cell.font = Font(color="0000FF")
        cell.fill = WARN_FILL

# Block 4: Health metrics
start = 33
ws.cell(row=start, column=1, value="БЛОК 4. Юнит-экономика — здоровье").font = HEAD
ws.cell(row=start, column=1).fill = HEAD_FILL

# ARPU is at row arpu_row=21, col B
metrics = [
    ("ARPU средневзв., ₽/мес", f"=B{arpu_row}"),
    ("Gross margin, %", 0.62),
    ("LTV (12 мес × ARPU × margin), ₽", f"=B{arpu_row}*12*B35"),
    ("LTV / CAC", f"=B36/B7"),
    ("Точка безубыточности, платящих клиентов", f"=B6/B{arpu_row}"),
    ("Точка безубыточности, MRR ₽/мес", f"=B6"),
    ("Целевой объём платящих к M18", 1500),
    ("Целевой MRR к M18, ₽", f"=B40*B{arpu_row}"),
]
for i, (k, v) in enumerate(metrics, start=start + 1):
    ws.cell(row=i, column=1, value=k).font = BOLD
    cell = ws.cell(row=i, column=2, value=v)
    cell.font = BOLD
    cell.fill = GOOD_FILL
    if isinstance(v, str) and v.startswith("=") and "%" not in k and "/" not in k:
        cell.number_format = '#,##0 ₽'

# Block 5: targets/checks
start = 43
ws.cell(row=start, column=1, value="БЛОК 5. Целевые ориентиры (для проверки)").font = HEAD
ws.cell(row=start, column=1).fill = HEAD_FILL
checks = [
    "• LTV/CAC ≥ 3.0 — здоровая SaaS-юнит-эконмика",
    "• Gross margin ≥ 60% — норма для SaaS",
    "• Точка безубыточности ≤ 1000 платящих — реально к M14-15",
    "• Доля LLM в COGS ≤ 30% выручки — иначе агрессивный кеш + Qwen self-hosted",
    "• Free → Paid conversion ≥ 10% (зрелая воронка)",
    "• ARPU 4 000+ ₽ к M15+ за счёт upsell на Pro/Enterprise",
    "• Net Dollar Retention ≥ 110% к M18",
]
for i, t in enumerate(checks, start=start + 1):
    ws.cell(row=i, column=1, value=t).font = THIN

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 18
for c in "CDEFG":
    ws.column_dimensions[c].width = 18


# -------------------------------------------------------------------- #
# Sheet 11 — Тарифные пакеты (5 тиров)                                  #
# -------------------------------------------------------------------- #
ws = wb["11. Тарифные пакеты"]
clear_sheet(ws)
ws["A1"] = "Omnia.AI — 11. Тарифные пакеты (v2, post-Promto: 5 уровней)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A2"] = "Цели: entry-tier ниже Promto (990 ₽ vs 690 ₽), но с mix LLM. Free как acquisition-engine. Pro/Ent с реальным backend для production."
ws["A2"].font = Font(italic=True, color="6B7280")
ws["A2"].alignment = WRAP

headers = ["Что входит", "Free", "Lite", "Starter", "Pro", "Enterprise"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

tiers_data = [
    ["Цена ₽/мес", "0", "990", "2 990", "7 990", "19 990"],
    ["Эквивалент Promto", "—", "Старт (690 ₽)", "Про (2 790 ₽)", "Про Макс (6 890 ₽)", "—"],
    ["Срок", "5 дней без карты", "месяц", "месяц", "месяц", "месяц или год -10%"],
    ["Кол-во проектов", "1", "1", "до 3", "до 10", "безлимит"],
    ["Кошелёк AI-токенов, ₽/мес", "500 (одноразово)", "1 000", "2 500", "6 000", "18 000"],
    ["Хостинг", "Поддомен omnia.ai/...", "Поддомен или один домен", "Кастом-домен .ru/.рф", "Выделенный VPS S", "Выделенный VPS M"],
    ["SSL", "Let's Encrypt", "Let's Encrypt", "Let's Encrypt", "Let's Encrypt + DV", "OV/EV по запросу"],
    ["Кастом-домен", "—", "1 включён", "1 включён", "до 2 доменов", "до 5 доменов"],
    ["Backend (БД + API)", "—", "—", "Postgres + REST + JWT", "+ Redis + S3", "+ кастом + DR-реплика"],
    ["Версии и rollback", "5 снапшотов", "20 снапшотов", "100 снапшотов", "500 + side-by-side", "безлимит"],
    ["Mix LLM", "Haiku + DeepSeek", "+ Gemini Flash", "+ Sonnet + GPT-4", "Все + кеш промптов", "Все + приоритет очереди"],
    ["Российские LLM (Yandex/GigaChat)", "—", "—", "да", "да", "★ default (152-ФЗ)"],
    ["Чат-боты (TG/VK)", "—", "—", "—", "★ да", "★ да + кастом"],
    ["Бизнес-автоматизации", "—", "—", "—", "★ да (n8n-style)", "★ да + кастом-сценарии"],
    ["1С-интеграция", "—", "—", "—", "—", "★ ДА"],
    ["GitHub-синк (опц.)", "—", "—", "—", "★ да", "★ да"],
    ["Self-export (ZIP/Docker)", "—", "ZIP", "ZIP", "ZIP + Docker", "ZIP + Docker"],
    ["Резервное копирование", "—", "—", "1×/сут, 7 дней", "1×/сут, 30 дней", "почасовое, 90 дней + офсайт"],
    ["Поддержка", "—", "Чат-бот", "Email SLA 48ч", "Email SLA 24ч", "Менеджер + SLA 4ч"],
    ["Часы инженеров (поверх подписки)", "—", "—", "5 000 ₽/час", "4 500 ₽/час", "4 000 ₽/час"],
    ["SLA доступности", "—", "—", "99.5%", "99.9%", "99.95% с фин.гарантиями"],
]

for r_i, row in enumerate(tiers_data, start=5):
    for c_i, val in enumerate(row, start=1):
        c = ws.cell(row=r_i, column=c_i, value=val)
        c.font = THIN
        c.alignment = WRAP
        if "★" in str(val):
            c.fill = GOOD_FILL

# Highlight Pro column as recommended
for r in range(4, 5 + len(tiers_data)):
    cell = ws.cell(row=r, column=5)
    if r == 4:
        cell.fill = PatternFill("solid", start_color="5E6AD2")
        cell.font = Font(bold=True, color="FFFFFF")
    elif not cell.fill or cell.fill.start_color.rgb != "00DCFCE7":
        cell.fill = ACCENT_FILL

ws.column_dimensions["A"].width = 38
for c in "BCDEF":
    ws.column_dimensions[c].width = 22

# add "Why each tier"
last = 5 + len(tiers_data) + 1
ws.cell(row=last + 1, column=1, value="ЛОГИКА ТАРИФОВ vs Promto").font = HEAD
ws.cell(row=last + 1, column=1).fill = HEAD_FILL
explainers = [
    ("Free", "Lead-magnet. Без карты. Цель — конверсия 10-12% во Lite в течение 30 дней."),
    ("Lite (990 ₽)", "Бьёт Promto Старт (690 ₽) ценой +43%, но даёт mix LLM (DeepSeek+Haiku+Gemini), кастом-домен, версионирование."),
    ("Starter (2 990 ₽)", "Паритет с Promto Про (2 790 ₽). Преимущество: реальный backend (Postgres/JWT), РФ-LLM."),
    ("Pro (7 990 ₽)", "Бьёт Promto Про Макс (6 890 ₽) ценой +16%, но даёт выделенный VPS, чат-боты, GitHub-синк, SLA."),
    ("Enterprise (19 990 ₽)", "Promto не имеет такого тира. Наш — 152-ФЗ + 1С + менеджер. Закрываем enterprise-сегмент полностью."),
]
for i, (k, v) in enumerate(explainers, start=last + 3):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.cell(row=i, column=2, value=v).alignment = WRAP


# -------------------------------------------------------------------- #
# Sheet 14 — GTM план (приоритеты vs Promto)                            #
# -------------------------------------------------------------------- #
ws = wb["14. GTM план"]
write_table(
    ws,
    "Omnia.AI — 14. GTM план (v2, фокус: обогнать Promto по платящим к M18)",
    "Стратегия: Free-funnel → Lite/Starter массой → upsell в Pro/Enterprise. Перехват Promto-трафика на низком тире.",
    ["Этап", "Цели по платящим", "Главный канал", "Подход", "Срок"],
    [
        ["Stage 0: Pre-launch", "500-1000 в waitlist", "Founder-led контент", "Manifesto-пост, демо-видео, статья «Чем Omnia отличается от Promto» (Habr+vc.ru). Pre-sales annual Lite по 5 940 ₽.", "M0–M3"],
        ["Stage 1: Free MVP", "300 Free signups, 50 Lite платных", "Yandex.Direct (Promto-альт. кеи) + Telegram", "Кеи: «promto альтернатива», «AI-сайт с откатом», «конструктор с backend». Бюджет Я.Директ 80к/мес.", "M4–M6"],
        ["Stage 2: Starter ramp", "200 платящих, ARPU 1 500", "Performance + контент-маркетинг", "Кейсы беттестеров. YouTube-серия Артёма «Собираем 5 продуктов на Omnia». Habr-статьи с сравнением.", "M7–M9"],
        ["Stage 3: Pro launch", "600 платящих, ARPU 2 900", "Перформ + рефералы + B2B2C", "Чат-боты как separate landing. Реферальная программа 200 ₽ обоим. Партнёрки с агентствами 30%-revenue-share.", "M10–M12"],
        ["Stage 4: Enterprise + 152-ФЗ", "1 000 платящих, ARPU 3 500", "Sales + sales-enablement", "Outbound к мидл-СМБ через личные письма. 1С-конференции. PR в Sostav/T-Ж.", "M13–M15"],
        ["Stage 5: Marketplace + scale", "1 500+ платящих, ARPU 4 200", "Виральные циклы + B2B2C", "Marketplace шаблонов, white-label для агентств, mobile app gen. NDR ≥ 110%.", "M16–M18"],
        ["", "", "", "", ""],
        ["─── ПЕРЕХВАТ PROMTO-ТРАФИКА ───", "", "", "", ""],
        ["1. Брендовые кеи в Я.Директ", "+30% воронки", "Y.Direct + VK Ads", "Ключи: «промто альтернатива», «promto vs», «promto отзывы», «конструктор с откатом версий». CTR-target 8%.", "M4+"],
        ["2. Сравнительные статьи", "+SEO-органика", "Habr + vc.ru + dev.to", "5 статей за 6 месяцев: «Omnia vs Promto», «Promto vs Tilda vs Omnia», «AI-конструкторы РФ 2026: обзор». Каждая со сравнительной таблицей и кейсом.", "M4–M9"],
        ["3. YouTube-обзоры с прямым сравнением", "+brand", "YouTube + TG", "«Делаю один и тот же сайт на Promto и Omnia — что выйдет быстрее/лучше». Цель — 50к просмотров на видео.", "M5+"],
        ["4. Интерактивный demo на лендинге", "+conv 3→5%", "Лендинг", "Уже сделано — story-driven демо с 5 версиями. Усилить выделенным разделом «vs Promto» с table.", "M0+"],
        ["5. Free trial без карты", "+30% Free signup", "Onboarding", "Та же фича что у Promto. Не упускать парирование.", "M3+"],
    ],
    widths=[44, 26, 26, 50, 14],
)


# -------------------------------------------------------------------- #
# Sheet 15 — Финмодель 18 мес (новый ARPU и объёмы)                     #
# -------------------------------------------------------------------- #
ws = wb["15. Финмодель 18 мес"]
clear_sheet(ws)
ws["A1"] = "Omnia.AI — 15. Финмодель 18 мес (v2, новая 5-tier модель, post-Promto)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A2"] = "ARPU растёт по мере зрелости когорты (Lite → Starter → Pro upsell). Объёмы платящих агрессивные, fueled Free-funnel'ом."
ws["A2"].font = Font(italic=True, color="6B7280")
ws["A2"].alignment = WRAP

# Months M1-M18
ws.cell(row=4, column=1, value="Параметр").font = HEAD
ws.cell(row=4, column=1).fill = HEAD_FILL
for m in range(1, 19):
    c = ws.cell(row=4, column=1 + m, value=f"M{m}")
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center")

# row data
def fill(row, label, values, fmt=None, fill_color=None):
    ws.cell(row=row, column=1, value=label).font = THIN
    for i, v in enumerate(values, start=2):
        c = ws.cell(row=row, column=i, value=v)
        c.font = THIN
        if fmt:
            c.number_format = fmt
        if fill_color:
            c.fill = fill_color

# Free signups
free_signups = [0, 0, 0, 200, 400, 700, 1100, 1600, 2300, 3200, 4400, 5800, 7400, 9200, 11200, 13500, 16000, 19000]
fill(5, "Free signups (новых, мес)", free_signups, fmt='#,##0')

# Paying (cumulative)
paying = [0, 0, 0, 5, 15, 35, 60, 120, 200, 320, 480, 680, 920, 1180, 1450, 1750, 2050, 2350]
fill(6, "Платящие клиенты (накопленно)", paying, fmt='#,##0')

# ARPU evolves: starts low (Lite-heavy) and grows
arpu = [990, 990, 990, 1100, 1200, 1400, 1700, 2000, 2300, 2600, 2850, 3050, 3250, 3450, 3650, 3850, 4000, 4080]
fill(7, "ARPU средневзв., ₽/мес", arpu, fmt='#,##0 ₽')

# MRR
ws.cell(row=8, column=1, value="MRR, ₽/мес").font = BOLD
for m in range(1, 19):
    c = ws.cell(row=8, column=1 + m, value=f"=B6*B7" if m == 1 else None)
    c.font = BOLD
    c.fill = ACCENT_FILL
    c.number_format = '#,##0 ₽'
    c.value = f"={openpyxl.utils.get_column_letter(1+m)}6*{openpyxl.utils.get_column_letter(1+m)}7"

# COGS LLM (≈25% of revenue)
ws.cell(row=9, column=1, value="COGS: LLM API (~25% MRR), ₽").font = THIN
for m in range(1, 19):
    c = ws.cell(row=9, column=1 + m,
                value=f"={openpyxl.utils.get_column_letter(1+m)}8*0.25")
    c.number_format = '#,##0 ₽'

# COGS servers (~10% of revenue)
ws.cell(row=10, column=1, value="COGS: серверы (~10% MRR), ₽").font = THIN
for m in range(1, 19):
    c = ws.cell(row=10, column=1 + m,
                value=f"={openpyxl.utils.get_column_letter(1+m)}8*0.10")
    c.number_format = '#,##0 ₽'

# Payment commissions ~3%
ws.cell(row=11, column=1, value="COGS: комиссии (~3% MRR), ₽").font = THIN
for m in range(1, 19):
    c = ws.cell(row=11, column=1 + m,
                value=f"={openpyxl.utils.get_column_letter(1+m)}8*0.03")
    c.number_format = '#,##0 ₽'

# Gross profit
ws.cell(row=12, column=1, value="ГРОСС-ПРИБЫЛЬ, ₽").font = BOLD
for m in range(1, 19):
    col = openpyxl.utils.get_column_letter(1 + m)
    c = ws.cell(row=12, column=1 + m,
                value=f"={col}8-{col}9-{col}10-{col}11")
    c.font = BOLD
    c.fill = GOOD_FILL
    c.number_format = '#,##0 ₽'

# OPEX FOT
fot = [0, 0, 0, 0, 0, 0, 100000, 100000, 100000, 280000, 280000, 280000, 500000, 500000, 1000000, 1000000, 1500000, 1500000]
fill(13, "OPEX: FOT команды, ₽", fot, fmt='#,##0 ₽')

# OPEX marketing
mkt = [10000, 10000, 10000, 80000, 80000, 80000, 100000, 100000, 100000, 200000, 200000, 250000, 300000, 350000, 400000, 500000, 600000, 700000]
fill(14, "OPEX: маркетинг, ₽", mkt, fmt='#,##0 ₽')

# OPEX office
office = [5000] * 8 + [10000] * 6 + [20000] * 4
fill(15, "OPEX: офис/юр/прочее, ₽", office, fmt='#,##0 ₽')

# EBITDA
ws.cell(row=16, column=1, value="EBITDA, ₽").font = BOLD
for m in range(1, 19):
    col = openpyxl.utils.get_column_letter(1 + m)
    c = ws.cell(row=16, column=1 + m,
                value=f"={col}12-{col}13-{col}14-{col}15")
    c.font = BOLD
    c.fill = WARN_FILL
    c.number_format = '#,##0 ₽'

# Cumulative cash
ws.cell(row=17, column=1, value="Накопленный кэш (со старта 30к ₽), ₽").font = BOLD
ws.cell(row=17, column=2, value="=30000+B16").number_format = '#,##0 ₽'
for m in range(2, 19):
    col = openpyxl.utils.get_column_letter(1 + m)
    prev = openpyxl.utils.get_column_letter(m)
    c = ws.cell(row=17, column=1 + m, value=f"={prev}17+{col}16")
    c.number_format = '#,##0 ₽'
    c.font = BOLD

# Summary
ws.cell(row=20, column=1, value="SUMMARY").font = HEAD
ws.cell(row=20, column=1).fill = HEAD_FILL

ws.cell(row=21, column=1, value="Накопленная выручка, ₽")
ws.cell(row=21, column=2, value="=SUM(B8:S8)").number_format = '#,##0 ₽'

ws.cell(row=22, column=1, value="Накопленная EBITDA, ₽").font = BOLD
ws.cell(row=22, column=2, value="=SUM(B16:S16)").number_format = '#,##0 ₽'
ws.cell(row=22, column=2).fill = GOOD_FILL

ws.cell(row=23, column=1, value="MRR на M18, ₽").font = BOLD
ws.cell(row=23, column=2, value="=S8").number_format = '#,##0 ₽'
ws.cell(row=23, column=2).fill = GOOD_FILL

ws.cell(row=24, column=1, value="ARR на M18, ₽").font = BOLD
ws.cell(row=24, column=2, value="=S8*12").number_format = '#,##0 ₽'
ws.cell(row=24, column=2).fill = GOOD_FILL

ws.cell(row=25, column=1, value="Платящих на M18")
ws.cell(row=25, column=2, value="=S6").number_format = '#,##0'

ws.cell(row=26, column=1, value="Месяц break-even MRR (когда EBITDA первый раз ≥ 0)")
ws.cell(row=26, column=2, value="M11–M12 (по плану)")

ws.column_dimensions["A"].width = 44
for col_letter in [openpyxl.utils.get_column_letter(c) for c in range(2, 20)]:
    ws.column_dimensions[col_letter].width = 13


# -------------------------------------------------------------------- #
# Sheet 18 — Bootstrap (новый break-even с 5 тарифами)                 #
# -------------------------------------------------------------------- #
ws = wb["18. Bootstrap-план"]
clear_sheet(ws)
ws["A1"] = "Omnia.AI — 18. Bootstrap-план (v2, новая 5-tier модель)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A2"] = "Старт с 30 000 ₽ (хостинг + API). Break-even MRR ожидаем на M11-M12. Free funnel радикально снижает CAC."
ws["A2"].font = Font(italic=True, color="6B7280")
ws["A2"].alignment = WRAP

blocks = [
    ("БЛОК 1. Стартовый капитал", []),
    ("Хостинг dev-сервера (SafeCloud S)", "3 000 ₽"),
    ("Домен omnia.ai (на год)", "2 000 ₽"),
    ("API-ключи LLM (DeepSeek + Haiku, 2 мес запас)", "10 000 ₽"),
    ("Dev-инфра (GitHub Pro, Figma, Notion)", "5 000 ₽"),
    ("Резерв", "10 000 ₽"),
    ("ИТОГО", "30 000 ₽"),
    ("", ""),
    ("БЛОК 2. Runway без выручки", []),
    ("M1-M3: 0₽ зарплаты, только инфра", "Burn ~5 000 ₽/мес. Учредители живут на текущий доход."),
    ("M4-M6: запуск Free + Lite", "Burn ~80 000 ₽/мес (маркетинг). Первая выручка к M5-M6."),
    ("M7-M9: Starter ramp", "Burn ~150 000 ₽/мес. Ожидаем MRR 200-400к."),
    ("M10-M11: Pro launch", "Burn ~580 000 ₽/мес (FOT 280к + mkt 200к + рост маркетинга). Ожидаем MRR 1M+."),
    ("M12+: Break-even", "MRR покрывает OPEX и начинает накапливать."),
    ("", ""),
    ("БЛОК 3. Killer-ходы (расширены)", []),
    ("1. Pre-sales annual Lite со скидкой 50%", "5 940 ₽ × 100 = 594 000 ₽ кэша к M3 + валидация рынка ДО кода."),
    ("2. Free trial как acquisition-engine", "10 000 Free → 1 200 платных к M9 при 12% conversion."),
    ("3. Промто-перехват в Я.Директ", "Кеи «промто альтернатива/promto vs/конструктор с откатом» = +30% воронки на старте."),
    ("4. Свои проекты как первые кейсы", "Артём перенесёт kanavto + kuzovnsk + klining-24 — реальные сайты с трафиком."),
    ("5. Telegram-сообщество AI-сайтострой РФ", "2 000 подписчиков к M6, 10 000 к M12. 2-4% конверсия = 200-400 платных."),
    ("6. B2B2C партнёрки 30% revenue-share", "5 партнёров × 5 клиентов/мес = 25 платных при нулевом CAC."),
    ("7. Outsourcing арбитраж кастом-разработки", "Pro/Ent заказывают 4-5к/час, мы платим студентам 1-1.5к. Маржа 70%."),
    ("8. Open-source LLM-роутер (бренд через DEV-комьюнити)", "Stars → traffic → клиенты. Эффект Lovable/Cursor."),
    ("", ""),
    ("БЛОК 4. Точка безубыточности (новая)", []),
    ("OPEX в стабильном состоянии (M12+), ₽/мес", "≈ 3 500 000"),
    ("ARPU средневзв. (зрелое M15+), ₽/мес", "≈ 4 080"),
    ("Break-even платящих", "≈ 858"),
    ("Целевой объём платящих к M18", "1 500+ → MRR 6.1М ₽ → ARR 73М ₽"),
    ("", ""),
    ("БЛОК 5. Когда брать ангела (если)", []),
    ("Условие 1: MRR ≥ 1М ₽ к M14-M15", "PMF подтверждён цифрами"),
    ("Условие 2: органический рост упёрся в потолок", "CAC растёт, нужен бренд + платная реклама"),
    ("Условие 3: гипотеза «5-10М ₽ ангел = 2.5x growth»", "Конкретный план как использовать капитал"),
    ("Сумма / доля", "5-10М ₽ за 10-12% (оценка post-money 50-80М)"),
    ("Без drag-along, advisory-share-tax, борд-сита", "Только инвестиция, без операционного контроля"),
]

for i, (k, v) in enumerate(blocks, start=4):
    if v == [] and isinstance(v, list):
        c = ws.cell(row=i, column=1, value=k)
        c.font = HEAD
        c.fill = HEAD_FILL
    else:
        ws.cell(row=i, column=1, value=k).font = THIN
        ws.cell(row=i, column=2, value=v).alignment = WRAP

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 60


wb.save(DST)
print(f"OK saved: {DST}")
print(f"size: {DST.stat().st_size} bytes")
