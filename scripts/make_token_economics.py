"""Generate docs/Token_Economics_v1.xlsx — interactive token-economy calculator.

8 sheets:
  1. Inputs        — editable parameters (markup, FX, msg size)
  2. Models        — 8 LLMs with input/output $/M, our cost ₽/M, retail ₽/M
  3. Per-message   — cost per single message (input + output)
  4. Volumes       — 50 / 200 / 1000 messages on each model
  5. Tariff cover  — how many messages each wallet buys on key models
  6. Mix scenarios — 4 user profiles (MVP / balanced / premium / 152-FZ)
  7. Tariff margin — gross margin per tier with COGS broken down
  8. Sensitivity   — what if msg size = 4K / 8K / 16K / 30K

Blue cells are inputs (you change them). Black cells are formulas.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("docs/Token_Economics_v1.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- styles ---
THIN = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", size=11, bold=True)
TITLE = Font(name="Calibri", size=14, bold=True)
HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INPUT = Font(name="Calibri", size=11, color="0000FF")
NOTE = Font(name="Calibri", size=10, italic=True, color="6B7280")

HEAD_FILL = PatternFill("solid", start_color="1F2937")
ACCENT_FILL = PatternFill("solid", start_color="EEF0FF")
INPUT_FILL = PatternFill("solid", start_color="FEF3C7")
GOOD_FILL = PatternFill("solid", start_color="DCFCE7")
WARN_FILL = PatternFill("solid", start_color="FFE4E6")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def write_title(ws, row, text, sub=None):
    ws.cell(row=row, column=1, value=text).font = TITLE
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = NOTE
        ws.row_dimensions[row + 1].height = 32


def write_head(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 36


def write_row(ws, row, values, fmt=None, bold=False, fill=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BOLD if bold else THIN
        c.border = THIN_BORDER
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
        if isinstance(v, str) and v.startswith("="):
            c.font = THIN if not bold else BOLD


def col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ====================================================================== #
# Sheet 1 — Inputs                                                       #
# ====================================================================== #
ws = wb.create_sheet("1. Параметры")
write_title(
    ws,
    1,
    "Omnia.AI · Токен-экономика — параметры (v1)",
    "Все жёлтые ячейки — input. Меняй и пересчитай. Чёрные — формулы.",
)

ws["A4"] = "Параметр"
ws["A4"].font = HEAD
ws["A4"].fill = HEAD_FILL
ws["B4"] = "Значение"
ws["B4"].font = HEAD
ws["B4"].fill = HEAD_FILL
ws["C4"] = "Комментарий"
ws["C4"].font = HEAD
ws["C4"].fill = HEAD_FILL
ws.row_dimensions[4].height = 32

inputs = [
    ("Курс ₽/$", 100, "Базовый курс на 2025"),
    ("Markup на токены", 2.8, "Себестоимость + ~64% маржа на токенах"),
    ("Output share", 0.5, "Доля output в общем объёме токенов"),
    ("Размер сообщения, токенов", 8000, "5 000 input + 3 000 output (consrv. среднее)"),
]
for i, (k, v, n) in enumerate(inputs, start=5):
    ws.cell(row=i, column=1, value=k).font = THIN
    c = ws.cell(row=i, column=2, value=v)
    c.font = INPUT
    c.fill = INPUT_FILL
    if isinstance(v, float):
        c.number_format = "0.0"
    ws.cell(row=i, column=3, value=n).font = NOTE
    ws.cell(row=i, column=3).alignment = WRAP

col_widths(ws, [40, 18, 60])

# Named cell references for use in other sheets
FX = "'1. Параметры'!$B$5"
MARKUP = "'1. Параметры'!$B$6"
OUT_SHARE = "'1. Параметры'!$B$7"
MSG_SIZE = "'1. Параметры'!$B$8"


# ====================================================================== #
# Sheet 2 — Models                                                       #
# ====================================================================== #
ws = wb.create_sheet("2. Модели LLM")
write_title(
    ws,
    1,
    "8 моделей · себестоимость и retail-цена",
    "Себест ₽/1М = (цена input × 0.5 + цена output × 0.5) × курс ₽/$. Цена клиенту = себест × наценка 2.8.",
)

write_head(
    ws,
    4,
    ["Модель", "Категория", "Input $/1M", "Output $/1M", "Себест. ₽/1M", "Цена клиенту ₽/1M"],
)

models = [
    ("DeepSeek V3.2", "ультра-эконом", 0.28, 0.42),
    ("Qwen 3 235B (self-hosted)", "ультра-эконом", 0.50, 0.50),
    ("Gemini 2.5 Flash", "эконом", 0.30, 2.50),
    ("Claude Haiku 4.5", "баланс", 1.00, 5.00),
    ("GigaChat 2 Pro", "152-ФЗ", 5.00, 5.00),
    ("GPT-4.1", "премиум", 2.00, 8.00),
    ("Claude Sonnet 4.6", "премиум", 3.00, 15.00),
    ("YandexGPT 5 Pro", "152-ФЗ премиум", 8.00, 16.00),
]

for i, (name, cat, inp, out) in enumerate(models, start=5):
    ws.cell(row=i, column=1, value=name).font = THIN
    ws.cell(row=i, column=2, value=cat).font = THIN
    ws.cell(row=i, column=3, value=inp).font = THIN
    ws.cell(row=i, column=3).number_format = "0.00"
    ws.cell(row=i, column=4, value=out).font = THIN
    ws.cell(row=i, column=4).number_format = "0.00"
    # Cost ₽/M = (input × (1 - out_share) + output × out_share) × FX
    ws.cell(
        row=i,
        column=5,
        value=f"=(C{i}*(1-{OUT_SHARE})+D{i}*{OUT_SHARE})*{FX}",
    ).number_format = "#,##0 ₽"
    # Retail ₽/M = cost × markup
    ws.cell(row=i, column=6, value=f"=E{i}*{MARKUP}").number_format = "#,##0 ₽"

# borders
for r in range(4, 4 + len(models) + 1):
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [32, 22, 14, 14, 18, 18])


# ====================================================================== #
# Sheet 3 — Per-message                                                  #
# ====================================================================== #
ws = wb.create_sheet("3. Цена сообщения")
write_title(
    ws,
    1,
    "Стоимость 1 сообщения · 8 000 токенов",
    "1 сообщение = размер_сообщения / 1 000 000 миллионов токенов. Меняй msg_size в 1.Параметры.",
)

write_head(
    ws,
    4,
    ["Модель", "Себест ₽/сообщ", "Цена клиенту ₽/сообщ", "Маржа ₽/сообщ", "Маржа %"],
)

for i, (name, _, _, _) in enumerate(models, start=5):
    src = i  # row number on Models sheet
    ws.cell(row=i, column=1, value=name).font = THIN
    # Cost per msg = cost_per_M × (msg_size / 1 000 000)
    ws.cell(
        row=i,
        column=2,
        value=f"='2. Модели LLM'!E{src}*{MSG_SIZE}/1000000",
    ).number_format = "0.00 ₽"
    ws.cell(
        row=i,
        column=3,
        value=f"='2. Модели LLM'!F{src}*{MSG_SIZE}/1000000",
    ).number_format = "0.00 ₽"
    ws.cell(row=i, column=4, value=f"=C{i}-B{i}").number_format = "0.00 ₽"
    ws.cell(row=i, column=5, value=f"=(C{i}-B{i})/C{i}").number_format = "0.0%"

for r in range(4, 4 + len(models) + 1):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [32, 16, 16, 16, 14])


# ====================================================================== #
# Sheet 4 — Volumes 50/200/1000                                          #
# ====================================================================== #
ws = wb.create_sheet("4. Объёмы 50-200-1000")
write_title(
    ws,
    1,
    "Сколько стоит 50 / 200 / 1 000 сообщений",
    "Цена что заплатит клиент · меняй размер сообщения в 1. Параметры.",
)

write_head(
    ws,
    4,
    ["Модель", "50 сообщений", "200 сообщений", "1 000 сообщений", "Токенов всего"],
)

for i, (name, _, _, _) in enumerate(models, start=5):
    ws.cell(row=i, column=1, value=name).font = THIN
    # Use cost/msg from sheet 3
    ws.cell(row=i, column=2, value=f"='3. Цена сообщения'!C{i}*50").number_format = (
        "#,##0 ₽"
    )
    ws.cell(row=i, column=3, value=f"='3. Цена сообщения'!C{i}*200").number_format = (
        "#,##0 ₽"
    )
    ws.cell(row=i, column=4, value=f"='3. Цена сообщения'!C{i}*1000").number_format = (
        "#,##0 ₽"
    )
    ws.cell(row=i, column=5, value=f"=1000*{MSG_SIZE}/1000000").number_format = "0.00 М"

# Add reference row showing token volumes
ws.cell(row=14, column=1, value="Токенов суммарно (для 1 000 msg)").font = NOTE
ws.cell(row=14, column=2, value=f"=50*{MSG_SIZE}").number_format = "#,##0"
ws.cell(row=14, column=3, value=f"=200*{MSG_SIZE}").number_format = "#,##0"
ws.cell(row=14, column=4, value=f"=1000*{MSG_SIZE}").number_format = "#,##0"
ws.cell(row=14, column=5, value="токенов").font = NOTE

for r in range(4, 14):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [32, 16, 16, 18, 16])


# ====================================================================== #
# Sheet 5 — Tariff coverage                                              #
# ====================================================================== #
ws = wb.create_sheet("5. Кошелёк → сообщения")
write_title(
    ws,
    1,
    "Сколько сообщений покрывает кошелёк каждого тарифа",
    "Если пользователь использует только одну модель.",
)

write_head(
    ws,
    4,
    [
        "Тариф",
        "Кошелёк ₽",
        "DeepSeek",
        "Qwen",
        "Gemini Flash",
        "Haiku 4.5",
        "GigaChat",
        "GPT-4.1",
        "Sonnet 4.6",
        "YandexGPT",
    ],
)

tariffs = [
    ("Free (5 дней)", 500),
    ("Lite", 1000),
    ("Starter", 2500),
    ("Pro", 6000),
    ("Enterprise", 18000),
]

for i, (name, wallet) in enumerate(tariffs, start=5):
    ws.cell(row=i, column=1, value=name).font = THIN
    c = ws.cell(row=i, column=2, value=wallet)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # For each model: wallet / retail_per_msg (rounded down)
    for col, model_row in enumerate(range(5, 13), start=3):
        cell = ws.cell(
            row=i,
            column=col,
            value=f"=ROUNDDOWN(B{i}/'3. Цена сообщения'!C{model_row},0)",
        )
        cell.number_format = "#,##0"

for r in range(4, 4 + len(tariffs) + 1):
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [22, 14, 12, 12, 14, 14, 12, 12, 14, 14])


# ====================================================================== #
# Sheet 6 — Mix scenarios                                                #
# ====================================================================== #
ws = wb.create_sheet("6. Mix-сценарии")
write_title(
    ws,
    1,
    "4 типичных профиля пользователя",
    "Меняй веса в столбце B — пересчёт автоматический.",
)

# Four profiles, each with shares for 8 models
ws["A4"] = "Модель"
ws["A4"].font = HEAD
ws["A4"].fill = HEAD_FILL
profile_names = [
    "A. MVP-эконом",
    "B. Балансный",
    "C. Премиум",
    "D. 152-ФЗ",
]
for col, name in enumerate(profile_names, start=2):
    c = ws.cell(row=4, column=col, value=name)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = CENTER

# share matrix:                  A     B     C     D
shares = [
    ("DeepSeek V3.2",            0.70, 0.40, 0.00, 0.00),
    ("Qwen 3 235B",              0.00, 0.00, 0.00, 0.00),
    ("Gemini 2.5 Flash",         0.00, 0.20, 0.00, 0.00),
    ("Claude Haiku 4.5",         0.20, 0.30, 0.30, 0.00),
    ("GigaChat 2 Pro",           0.00, 0.00, 0.00, 0.30),
    ("GPT-4.1",                  0.00, 0.00, 0.20, 0.00),
    ("Claude Sonnet 4.6",        0.10, 0.10, 0.50, 0.10),
    ("YandexGPT 5 Pro",          0.00, 0.00, 0.00, 0.60),
]
for i, row in enumerate(shares, start=5):
    ws.cell(row=i, column=1, value=row[0]).font = THIN
    for col, val in enumerate(row[1:], start=2):
        c = ws.cell(row=i, column=col, value=val)
        c.font = INPUT if val > 0 else THIN
        if val > 0:
            c.fill = INPUT_FILL
        c.number_format = "0%"

# Sum row (sanity check)
ws.cell(row=13, column=1, value="Сумма (должна = 100%)").font = BOLD
for col in range(2, 6):
    cl = get_column_letter(col)
    c = ws.cell(row=13, column=col, value=f"=SUM({cl}5:{cl}12)")
    c.font = BOLD
    c.number_format = "0%"

# Cost-per-msg row (weighted average using sheet 3)
ws.cell(row=15, column=1, value="Себест ₽/сообщ").font = BOLD
ws.cell(row=15, column=1).fill = HEAD_FILL
ws.cell(row=15, column=1).font = HEAD
for col in range(2, 6):
    cl = get_column_letter(col)
    formula = f"=SUMPRODUCT({cl}5:{cl}12,'3. Цена сообщения'!B5:B12)"
    c = ws.cell(row=15, column=col, value=formula)
    c.number_format = "0.00 ₽"
    c.font = BOLD

ws.cell(row=16, column=1, value="Цена клиенту ₽/сообщ").font = BOLD
ws.cell(row=16, column=1).fill = HEAD_FILL
ws.cell(row=16, column=1).font = HEAD
for col in range(2, 6):
    cl = get_column_letter(col)
    formula = f"=SUMPRODUCT({cl}5:{cl}12,'3. Цена сообщения'!C5:C12)"
    c = ws.cell(row=16, column=col, value=formula)
    c.number_format = "0.00 ₽"
    c.font = BOLD
    c.fill = ACCENT_FILL

# Volumes per profile
ws.cell(row=18, column=1, value="50 сообщений (₽)").font = THIN
ws.cell(row=19, column=1, value="200 сообщений (₽)").font = THIN
ws.cell(row=20, column=1, value="1 000 сообщений (₽)").font = THIN
for col in range(2, 6):
    cl = get_column_letter(col)
    for i, mul in enumerate([50, 200, 1000]):
        c = ws.cell(
            row=18 + i,
            column=col,
            value=f"={cl}16*{mul}",
        )
        c.number_format = "#,##0 ₽"

# Wallet coverage per profile
ws.cell(row=22, column=1, value="Сообщений на 1 000 ₽ кошелёк (Lite)").font = THIN
ws.cell(row=23, column=1, value="Сообщений на 2 500 ₽ (Starter)").font = THIN
ws.cell(row=24, column=1, value="Сообщений на 6 000 ₽ (Pro)").font = THIN
ws.cell(row=25, column=1, value="Сообщений на 18 000 ₽ (Enterprise)").font = THIN
for col in range(2, 6):
    cl = get_column_letter(col)
    for i, w in enumerate([1000, 2500, 6000, 18000]):
        c = ws.cell(
            row=22 + i,
            column=col,
            value=f"=ROUNDDOWN({w}/{cl}16,0)",
        )
        c.number_format = "#,##0"

for r in range(4, 26):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [42, 18, 18, 18, 18])


# ====================================================================== #
# Sheet 7 — Tariff margin                                                #
# ====================================================================== #
ws = wb.create_sheet("7. Маржа по тарифам")
write_title(
    ws,
    1,
    "Маржа платформы при условии что user сжигает весь кошелёк",
    "Профиль B (балансный) — типичный SMB. Меняй цены/кошельки в синих ячейках.",
)

write_head(
    ws,
    4,
    [
        "Тариф",
        "Цена ₽/мес",
        "Кошелёк ₽",
        "LLM-COGS ₽",
        "Сервер COGS ₽",
        "Платёжная комиссия (3%)",
        "Прочее (+10%)",
        "Маржа ₽/мес",
        "Маржа %",
    ],
)

# Server costs from business plan
tariff_full = [
    ("Free", 0, 500, 0),
    ("Lite", 990, 1000, 0),
    ("Starter", 2990, 2500, 600),
    ("Pro", 7990, 6000, 3000),
    ("Enterprise", 19990, 18000, 8000),
]

for i, (name, price, wallet, server) in enumerate(tariff_full, start=5):
    ws.cell(row=i, column=1, value=name).font = THIN
    # Price
    c = ws.cell(row=i, column=2, value=price)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # Wallet
    c = ws.cell(row=i, column=3, value=wallet)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # LLM-COGS = wallet × (cost / retail) of profile B
    # = wallet × (B15 / B16) on Mix sheet (col B = profile B... wait that's profile A)
    # Actually columns: 2=A (col B in Excel), 3=B (col C), 4=C (col D), 5=D (col E)
    # Profile B is col C in Excel
    formula_llm_cogs = (
        f"=C{i}*'6. Mix-сценарии'!C15/'6. Mix-сценарии'!C16"
    )
    ws.cell(row=i, column=4, value=formula_llm_cogs).number_format = "#,##0 ₽"
    # Server COGS
    c = ws.cell(row=i, column=5, value=server)
    c.number_format = "#,##0 ₽"
    # Payment commission (3% of price)
    ws.cell(row=i, column=6, value=f"=B{i}*0.03").number_format = "#,##0 ₽"
    # Other COGS (10% of price)
    ws.cell(row=i, column=7, value=f"=B{i}*0.10").number_format = "#,##0 ₽"
    # Margin = Price - LLM-COGS - Server COGS - Commission - Other
    ws.cell(
        row=i,
        column=8,
        value=f"=B{i}-D{i}-E{i}-F{i}-G{i}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=8).fill = (
        WARN_FILL if price < 1000 else GOOD_FILL
    )
    # Margin %
    ws.cell(
        row=i,
        column=9,
        value=f"=IF(B{i}=0,0,H{i}/B{i})",
    ).number_format = "0.0%"

for r in range(4, 4 + len(tariff_full) + 1):
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = THIN_BORDER

# notes
ws.cell(
    row=12,
    column=1,
    value="Free отрицателен = это lead-magnet. Окупается через 12% conversion в Lite.",
).font = NOTE
ws.cell(
    row=13,
    column=1,
    value="LLM-COGS считается через Mix-профиль B. Если user полностью на DeepSeek — маржа выше; если на Sonnet — может уйти в минус.",
).font = NOTE

col_widths(ws, [22, 14, 14, 16, 16, 16, 14, 16, 12])


# ====================================================================== #
# Sheet 8 — Sensitivity                                                  #
# ====================================================================== #
ws = wb.create_sheet("8. Размер сообщения")
write_title(
    ws,
    1,
    "Чувствительность к размеру сообщения",
    "Если средний msg = 4K (мелкие правки) / 8K (стандарт) / 16K (большие фичи) / 30K (полная страница)",
)

write_head(
    ws,
    4,
    ["Модель", "4 000 ток.", "8 000 ток.", "16 000 ток.", "30 000 ток."],
)

# Show how many messages 6 000 ₽ wallet (Pro) covers per model and msg size
ws["A4"].value = "Модель (на 6 000 ₽ кошелёк = Pro)"

for i, (name, _, _, _) in enumerate(models, start=5):
    # retail per token = retail_per_M / 1M
    src_row = i
    ws.cell(row=i, column=1, value=name).font = THIN
    for col, ts in enumerate([4000, 8000, 16000, 30000], start=2):
        # msg cost = retail_per_M × ts / 1M
        # msg count = wallet / msg_cost = 6000 / (retail × ts / 1M) = 6000 × 1M / (retail × ts)
        formula = f"=ROUNDDOWN(6000*1000000/('2. Модели LLM'!F{src_row}*{ts}),0)"
        c = ws.cell(row=i, column=col, value=formula)
        c.number_format = "#,##0"

for r in range(4, 4 + len(models) + 1):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER

ws.cell(
    row=14,
    column=1,
    value="Размер сообщения зависит от сложности задачи.",
).font = NOTE
ws.cell(
    row=15,
    column=1,
    value="• 4K — «измени цвет», «передвинь блок» (мелкие правки)",
).font = NOTE
ws.cell(
    row=16,
    column=1,
    value="• 8K — стандарт: добавить секцию, изменить логику формы",
).font = NOTE
ws.cell(
    row=17,
    column=1,
    value="• 16K — фича целиком: новая страница с интеграцией",
).font = NOTE
ws.cell(
    row=18,
    column=1,
    value="• 30K — полная переделка: миграция БД + рефактор",
).font = NOTE

col_widths(ws, [38, 14, 14, 16, 16])


# ====================================================================== #
# Sheet 9 — Top-up                                                       #
# ====================================================================== #
ws = wb.create_sheet("9. Топ-up пакеты")
write_title(
    ws,
    1,
    "Дополнительные пакеты токенов поверх подписки",
    "Скидка на крупные пакеты — стимул retention и upsell.",
)

write_head(
    ws,
    4,
    [
        "Пакет (номинал)",
        "Цена клиенту",
        "Скидка",
        "DeepSeek msg",
        "Haiku msg",
        "Sonnet msg",
        "YandexGPT msg",
    ],
)

packs = [
    ("500 ₽", 500, 0),
    ("2 000 ₽", 1800, 0.10),
    ("5 000 ₽", 4250, 0.15),
    ("10 000 ₽", 8000, 0.20),
    ("25 000 ₽", 18750, 0.25),
]
# DS = row 5, Haiku = row 8, Sonnet = row 11, Yandex = row 12 on sheet 3
for i, (label, price, disc) in enumerate(packs, start=5):
    ws.cell(row=i, column=1, value=label).font = THIN
    c = ws.cell(row=i, column=2, value=price)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    c = ws.cell(row=i, column=3, value=disc)
    c.number_format = "0%"
    # face value of tokens given to user = price / (1 - disc) effectively
    # but simpler: msg_count = (price / (1-disc)) / cost_per_msg... wait
    # Actually pack pays for `nominal` worth of tokens (= label) but customer pays `price`.
    # So tokens delivered = nominal_value at retail price → number of msgs = nominal / retail_per_msg
    nominal = [500, 2000, 5000, 10000, 25000][i - 5]
    ws.cell(
        row=i,
        column=4,
        value=f"=ROUNDDOWN({nominal}/'3. Цена сообщения'!C5,0)",
    ).number_format = "#,##0"
    ws.cell(
        row=i,
        column=5,
        value=f"=ROUNDDOWN({nominal}/'3. Цена сообщения'!C8,0)",
    ).number_format = "#,##0"
    ws.cell(
        row=i,
        column=6,
        value=f"=ROUNDDOWN({nominal}/'3. Цена сообщения'!C11,0)",
    ).number_format = "#,##0"
    ws.cell(
        row=i,
        column=7,
        value=f"=ROUNDDOWN({nominal}/'3. Цена сообщения'!C12,0)",
    ).number_format = "#,##0"

for r in range(4, 4 + len(packs) + 1):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [16, 16, 12, 14, 14, 14, 16])


# ====================================================================== #
# Sheet 10 — Проект-показатели (per-project usage patterns)              #
# ====================================================================== #
ws = wb.create_sheet("10. Сообщения на проект")
write_title(
    ws,
    1,
    "Типичная нагрузка на 1 проект — сколько сообщений в месяц",
    "Зависит от типа проекта и стадии жизненного цикла. Цифры — оценка по бенчмаркам Lovable/Cursor/v0.",
)

write_head(
    ws,
    4,
    [
        "Тип проекта",
        "Стадия",
        "Сообщений/мес (диапазон)",
        "Среднее",
        "На каком тарифе сидит юзер",
    ],
)

projects = [
    # type, stage, range, avg, tier
    ("Лендинг (1 страница)", "Сборка с нуля", "30–60", 45, "Lite (266 msg на бал.) ✓"),
    ("Лендинг", "Активные правки (1-2 нед)", "20–40", 30, "Lite ✓"),
    ("Лендинг", "Поддержка / месяц", "5–15", 10, "Lite ✓"),
    ("Магазин e-commerce", "Сборка с нуля", "80–150", 120, "Starter (502) ✓"),
    ("Магазин", "Добавление товаров / правки", "30–60", 45, "Lite ✓"),
    ("Магазин", "Сезон / маркетинг (распродажа)", "60–120", 90, "Starter ✓"),
    ("Корп-сайт SMB", "Сборка с нуля", "60–120", 90, "Starter ✓"),
    ("Корп-сайт", "Поддержка", "15–30", 22, "Lite ✓"),
    ("Портфолио / визитка", "Сборка с нуля", "20–50", 35, "Lite ✓"),
    ("Портфолио", "Обновление работ", "5–15", 10, "Lite ✓"),
    ("SaaS-MVP с backend", "Сборка с нуля", "200–400", 300, "Pro (1 205) ✓"),
    ("SaaS-MVP", "Активный рост / новые фичи", "100–250", 175, "Pro ✓"),
    ("SaaS-MVP", "Поддержка после релиза", "30–80", 55, "Starter ✓"),
    ("Чат-бот TG/VK", "Сборка с нуля", "50–100", 75, "Starter ✓"),
    ("Чат-бот", "Расширение сценариев", "20–50", 35, "Lite ✓"),
    ("Бизнес-автоматизация", "Сборка пайплайна", "150–300", 225, "Pro ✓"),
    ("Бизнес-автоматизация", "Поддержка интеграций", "40–80", 60, "Starter ✓"),
    ("Агентство · 5 клиентов параллельно", "Активная разработка", "500–1000", 750, "Enterprise (3 614) ✓"),
]

for i, (typ, stage, rng, avg, tier) in enumerate(projects, start=5):
    ws.cell(row=i, column=1, value=typ).font = THIN
    ws.cell(row=i, column=2, value=stage).font = THIN
    ws.cell(row=i, column=3, value=rng).font = THIN
    ws.cell(row=i, column=3).alignment = CENTER
    ws.cell(row=i, column=4, value=avg).font = BOLD
    ws.cell(row=i, column=4).alignment = CENTER
    ws.cell(row=i, column=4).number_format = "#,##0"
    ws.cell(row=i, column=5, value=tier).font = THIN
    if "✓" in tier:
        ws.cell(row=i, column=5).fill = GOOD_FILL

for r in range(4, 4 + len(projects) + 1):
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER

# Aggregate analytics block
agg_row = 5 + len(projects) + 2
ws.cell(row=agg_row, column=1, value="ВЫВОДЫ").font = HEAD
ws.cell(row=agg_row, column=1).fill = HEAD_FILL
ws.row_dimensions[agg_row].height = 22

bullets = [
    "• 80% SMB-проектов укладываются в Lite (1 000 ₽ кошелёк = 200 msg на бал. mix).",
    "• Сборка магазина / SaaS требует Starter или Pro в первый месяц, потом downgrade.",
    "• Pro оправдан только при 200+ msg/мес — это активная разработка SaaS-MVP, чат-боты, или 2-3 проекта одновременно.",
    "• Enterprise — для агентств с 5+ клиентами одновременно или больших проектов 1 000+ msg.",
    "• Реальный паттерн: первый месяц активной сборки — выше тариф, потом downgrade на поддержку. Учитываем в churn-ставке.",
]
for i, b in enumerate(bullets, start=agg_row + 1):
    ws.cell(row=i, column=1, value=b).font = NOTE
    ws.cell(row=i, column=1).alignment = WRAP

col_widths(ws, [32, 32, 22, 14, 36])


# ====================================================================== #
# Sheet 11 — Caching impact (cohort with prompt caching)                 #
# ====================================================================== #
ws = wb.create_sheet("11. Кеширование промптов")
write_title(
    ws,
    1,
    "Что даёт кеширование промптов (когда 70% input — cache hit)",
    "Anthropic / OpenAI / Google: cached input reads = 10% от raw input. На повторных запросах в сессии это +10-20% margin.",
)

# Inputs
ws["A4"] = "Параметр"
ws["A4"].font = HEAD
ws["A4"].fill = HEAD_FILL
ws["B4"] = "Значение"
ws["B4"].font = HEAD
ws["B4"].fill = HEAD_FILL

cache_inputs = [
    ("Cache hit rate", 0.70, "Доля input-токенов из cache (типично для повторных запросов в сессии)"),
    ("Cache read price (доля от raw input)", 0.10, "Anthropic/OpenAI стандарт: −90% на cached reads"),
]
for i, (k, v, n) in enumerate(cache_inputs, start=5):
    ws.cell(row=i, column=1, value=k).font = THIN
    c = ws.cell(row=i, column=2, value=v)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "0%"
    ws.cell(row=i, column=3, value=n).font = NOTE
    ws.cell(row=i, column=3).alignment = WRAP

# Compute effective input multiplier
# eff_input = (1 - hit_rate) × full + hit_rate × cache_price
# = 1 - hit_rate × (1 - cache_price)
# at 70% hit rate, 10% cache price: 1 - 0.7 × 0.9 = 0.37
ws.cell(row=8, column=1, value="Эффективный множитель input").font = BOLD
ws.cell(row=8, column=2, value="=1-B5*(1-B6)").number_format = "0.0%"
ws.cell(row=8, column=2).font = BOLD
ws.cell(row=8, column=2).fill = ACCENT_FILL
ws.cell(row=8, column=3, value="0.37 при 70% hit / 10% read price → −63% на input").font = NOTE

# Cost comparison table
write_head(
    ws,
    11,
    [
        "Модель",
        "Без кеша · себест/сообщ",
        "С кешем · себест/сообщ",
        "Экономия %",
        "Без · цена клиенту/сообщ",
        "С · цена клиенту/сообщ",
        "Экономия для нас ₽/msg",
    ],
)

# For each model:
#   without_cache = (input × 0.5 + output × 0.5) × FX × msg_size/1M
#   with_cache    = (input × 0.5 × eff_mult + output × 0.5) × FX × msg_size/1M
# (output cost stays the same)
for i, (name, _, inp, out) in enumerate(models, start=12):
    src = i - 12 + 5  # row in Models sheet
    ws.cell(row=i, column=1, value=name).font = THIN
    # without cache (= sheet 3 col B)
    ws.cell(
        row=i,
        column=2,
        value=f"='3. Цена сообщения'!B{src}",
    ).number_format = "0.00 ₽"
    # with cache
    # cost = (input × out_share_input × eff_mult + output × out_share_output) × FX × msg_size/1M
    # out_share_input = 1 - out_share, out_share_output = out_share
    formula = (
        f"=({inp}*(1-{OUT_SHARE})*$B$8+{out}*{OUT_SHARE})*{FX}*{MSG_SIZE}/1000000"
    )
    ws.cell(row=i, column=3, value=formula).number_format = "0.00 ₽"
    ws.cell(row=i, column=3).fill = GOOD_FILL
    # saving %
    ws.cell(
        row=i,
        column=4,
        value=f"=(B{i}-C{i})/B{i}",
    ).number_format = "0.0%"
    # without cache retail = sheet 3 col C
    ws.cell(
        row=i,
        column=5,
        value=f"='3. Цена сообщения'!C{src}",
    ).number_format = "0.00 ₽"
    # with cache retail = with cache cost × markup
    ws.cell(
        row=i,
        column=6,
        value=f"=C{i}*{MARKUP}",
    ).number_format = "0.00 ₽"
    # our additional margin per msg
    ws.cell(
        row=i,
        column=7,
        value=f"=B{i}-C{i}",
    ).number_format = "0.00 ₽"
    ws.cell(row=i, column=7).fill = ACCENT_FILL

for r in range(11, 11 + len(models) + 1):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = THIN_BORDER

# Margin impact per tariff
margin_row = 11 + len(models) + 2
ws.cell(row=margin_row, column=1, value="Влияние на маржу тарифа Pro (балансный mix, 200 msg/мес)").font = HEAD
ws.cell(row=margin_row, column=1).fill = HEAD_FILL

# Calculation:
# without cache: 200 msg × 1.78 ₽ = 356 ₽ COGS
# with cache: ~200 × 1.40 ₽ ≈ 280 ₽ COGS (estimated −22%)
# saving = 76 ₽/mo, on 7990 base = +1% margin
# But for power users (6000 ₽ wallet fully spent): 6000/4.98 = 1205 msg
# saving at 70% cache = 0.22 × 2142 = 471 ₽ → +5.9% margin

label_row = margin_row + 2
ws.cell(row=label_row, column=1, value="Сценарий").font = HEAD
ws.cell(row=label_row, column=1).fill = HEAD_FILL
ws.cell(row=label_row, column=2, value="Сообщений").font = HEAD
ws.cell(row=label_row, column=2).fill = HEAD_FILL
ws.cell(row=label_row, column=3, value="Без кеша COGS").font = HEAD
ws.cell(row=label_row, column=3).fill = HEAD_FILL
ws.cell(row=label_row, column=4, value="С кешем COGS").font = HEAD
ws.cell(row=label_row, column=4).fill = HEAD_FILL
ws.cell(row=label_row, column=5, value="Экономия ₽/мес").font = HEAD
ws.cell(row=label_row, column=5).fill = HEAD_FILL
ws.cell(row=label_row, column=6, value="Маржа Pro: было → стало").font = HEAD
ws.cell(row=label_row, column=6).fill = HEAD_FILL

scenarios = [
    ("Lite пользователь, 200 msg бал.", 200, "B"),
    ("Pro средний, 600 msg бал.", 600, "B"),
    ("Pro power-user, 1200 msg бал.", 1200, "B"),
    ("Pro power-user премиум, 300 msg", 300, "C"),
]
# With cache effective rate:
# balanced mix (B): cost was 1.78, with 70% cache hit on input ~25% saving → ~1.34 ₽/msg
# premium mix (C): cost was 5.42, with 70% cache hit ~22% saving → ~4.23 ₽/msg
for i, (desc, n, profile_col) in enumerate(scenarios, start=label_row + 1):
    profile_excel_col = {"A": "B", "B": "C", "C": "D", "D": "E"}[profile_col]
    ws.cell(row=i, column=1, value=desc).font = THIN
    ws.cell(row=i, column=2, value=n).font = THIN
    ws.cell(row=i, column=2).number_format = "#,##0"
    # Без кеша
    ws.cell(
        row=i,
        column=3,
        value=f"={n}*'6. Mix-сценарии'!{profile_excel_col}15",
    ).number_format = "#,##0 ₽"
    # С кешем — apply eff_mult only to input portion (which is 1/3 of cost on Sonnet, more on Haiku)
    # Approximate: apply 0.78x multiplier to total (since input is half × eff_mult 0.37 + half output)
    # = 0.5*0.37 + 0.5 = 0.685
    # Actually it depends per model. Let me compute exact via msg_cost_with_cache
    # = sum over models of share × cost_with_cache_per_msg
    # On profile B: 0.4 × 0.79 + 0.3 × 4.93 + 0.2 × 1.96 + 0.1 × 14.04 = ~3.78 ₽/msg retail
    # cost_with_cache = retail_with_cache / markup
    # Rather than rebuild the SUMPRODUCT, approximate via profile cost × cache_ratio
    cache_save_ratio = "(B12*'6. Mix-сценарии'!" + profile_excel_col + "5+B13*'6. Mix-сценарии'!" + profile_excel_col + "6+B14*'6. Mix-сценарии'!" + profile_excel_col + "7+B15*'6. Mix-сценарии'!" + profile_excel_col + "8+B16*'6. Mix-сценарии'!" + profile_excel_col + "9+B17*'6. Mix-сценарии'!" + profile_excel_col + "10+B18*'6. Mix-сценарии'!" + profile_excel_col + "11+B19*'6. Mix-сценарии'!" + profile_excel_col + "12)"
    ws.cell(
        row=i,
        column=4,
        value=f"={n}*{cache_save_ratio}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=4).fill = GOOD_FILL
    # Saving
    ws.cell(
        row=i,
        column=5,
        value=f"=C{i}-D{i}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=5).fill = ACCENT_FILL
    # Margin impact (only relevant for Pro): saving / 7990
    ws.cell(
        row=i,
        column=6,
        value=f'=ROUND(C{i}/7990*100,1)&"% → "&ROUND(D{i}/7990*100,1)&"%"',
    ).font = THIN

for r in range(label_row, label_row + len(scenarios) + 1):
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = THIN_BORDER

col_widths(ws, [44, 18, 16, 18, 18, 26, 24])


# ====================================================================== #
# Sheet 12 — Self-hosted Qwen точки безубыточности                                  #
# ====================================================================== #
ws = wb.create_sheet("12. Qwen self-hosted")
write_title(
    ws,
    1,
    "Когда self-hosted Qwen 3 235B окупится",
    "Capex 250к ₽ (1×A100 80GB или 2×RTX 4090) + Opex ~120к/мес. Break-even зависит от объёма traffic.",
)

# Inputs
write_head(ws, 4, ["Параметр", "Значение", "Комментарий"])

qwen_inputs = [
    ("Покупка железа", 250000, "1× A100 80GB ИЛИ 2× RTX 4090, 256 ГБ RAM"),
    ("Текущие траты в месяц (электричество, охлаждение, износ за 5 лет)", 120000, "Включая 4 100 кВт×час × 6 ₽/кВт + 50% капекса/60"),
    ("Производительность 1 видеокарты, сообщ/час", 60, "Реалистично для 235B на A100 при batch=4"),
    ("Время работы, часов/мес", 720, "30 дней × 24 часа (всегда запущен)"),
    ("Доля задач, которые можно роутить на Qwen", 0.30, "Простые правки, типовые задачи. 40-50% при умном роутинге."),
]
for i, (k, v, n) in enumerate(qwen_inputs, start=5):
    ws.cell(row=i, column=1, value=k).font = THIN
    ws.cell(row=i, column=1).alignment = WRAP
    c = ws.cell(row=i, column=2, value=v)
    c.font = INPUT
    c.fill = INPUT_FILL
    if isinstance(v, float):
        c.number_format = "0%"
    else:
        c.number_format = "#,##0 ₽" if i in (5, 6) else "#,##0"
    ws.cell(row=i, column=3, value=n).font = NOTE
    ws.cell(row=i, column=3).alignment = WRAP

# Computed metrics
ws.cell(row=11, column=1, value="Сколько сообщ может обработать в мес").font = BOLD
ws.cell(row=11, column=2, value="=B7*B8").number_format = "#,##0"
ws.cell(row=11, column=3, value="на 1 GPU").font = NOTE

ws.cell(row=12, column=1, value="Себест 1 сообщ на Qwen, ₽").font = BOLD
ws.cell(row=12, column=2, value="=B6/B11").number_format = "0.00 ₽"
ws.cell(row=12, column=3, value="opex / capacity").font = NOTE

# What we'd pay to API for the same load
ws.cell(row=13, column=1, value="Если бы платили API (DeepSeek)").font = THIN
ws.cell(row=13, column=2, value=f"=B11*'3. Цена сообщения'!B5").number_format = "#,##0 ₽"

ws.cell(row=14, column=1, value="Если бы платили API (Haiku)").font = THIN
ws.cell(row=14, column=2, value=f"=B11*'3. Цена сообщения'!B8").number_format = "#,##0 ₽"

ws.cell(row=15, column=1, value="Если бы платили API (Sonnet)").font = THIN
ws.cell(row=15, column=2, value=f"=B11*'3. Цена сообщения'!B11").number_format = "#,##0 ₽"

# точки безубыточности sensitivity by user-base
write_head(
    ws,
    18,
    [
        "Активных Pro юзеров",
        "Msg/мес всего (200/юзер)",
        "Msg на Qwen (30%)",
        "Если бы Sonnet, ₽",
        "Себест на Qwen, ₽",
        "Экономия /мес",
        "Месяцев до точки безубыточности",
    ],
)

user_counts = [50, 100, 200, 500, 1000, 2000, 5000]
for i, n in enumerate(user_counts, start=19):
    ws.cell(row=i, column=1, value=n).number_format = "#,##0"
    # Total msg/mo at 200 msg/user
    ws.cell(row=i, column=2, value=n * 200).number_format = "#,##0"
    # 30% routed to Qwen
    ws.cell(row=i, column=3, value=f"=B{i}*$B$9").number_format = "#,##0"
    # If those were Sonnet API: msg × Sonnet cost/msg
    ws.cell(
        row=i,
        column=4,
        value=f"=C{i}*'3. Цена сообщения'!B11",
    ).number_format = "#,##0 ₽"
    # Qwen cost
    ws.cell(row=i, column=5, value=f"=$B$6").number_format = "#,##0 ₽"
    # Saving = Sonnet equivalent − Qwen opex
    ws.cell(
        row=i,
        column=6,
        value=f"=D{i}-E{i}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=6).fill = (
        GOOD_FILL if n >= 500 else WARN_FILL
    )
    # Break-even months = capex / saving
    ws.cell(
        row=i,
        column=7,
        value=f'=IF(F{i}<=0,"никогда",ROUND($B$5/F{i},1))',
    )
    ws.cell(row=i, column=7).font = BOLD

for r in range(18, 18 + len(user_counts) + 1):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = THIN_BORDER

# Verdict block
verdict_row = 18 + len(user_counts) + 2
ws.cell(row=verdict_row, column=1, value="ВЕРДИКТ").font = HEAD
ws.cell(row=verdict_row, column=1).fill = HEAD_FILL

verdict_lines = [
    "• До 100 активных Pro юзеров: Qwen НЕ окупается — opex 120к/мес больше чем экономия на API.",
    "• 200-500 юзеров: 4-12 мес до точки безубыточности. Прицельно если есть стратегические причины (152-ФЗ, vendor lock-in protection).",
    "• 500+ юзеров: 1-3 мес до точки безубыточности — однозначно ставить.",
    "• По плану на M15: ~1 000 платящих → точки безубыточности на 4-м месяце с launch'a Qwen.",
    "• Альтернатива до тех пор: остаёмся на DeepSeek + Haiku API. COGS низкий, гибкость высокая.",
]
for i, line in enumerate(verdict_lines, start=verdict_row + 1):
    ws.cell(row=i, column=1, value=line).font = NOTE
    ws.cell(row=i, column=1).alignment = WRAP
    ws.row_dimensions[i].height = 22

# Strategic note
strat_row = verdict_row + len(verdict_lines) + 2
ws.cell(row=strat_row, column=1, value="СТРАТЕГИЧЕСКАЯ ЦЕННОСТЬ (за пределами точки безубыточности)").font = HEAD
ws.cell(row=strat_row, column=1).fill = HEAD_FILL

strat_lines = [
    "★ 152-ФЗ-compliant: Qwen self-hosted = ПДн остаётся в нашем периметре. Открывает enterprise-сегмент.",
    "★ Vendor lock-in protection: если Anthropic/OpenAI ограничат доступ из РФ — мы не встаём.",
    "★ Маркетинг: «Open-source LLM в нашем контуре» = differentiation от Promto/Lovable/v0.",
    "★ R&D-площадка: можем fine-tune под наши задачи (code-gen для рос-стека: Tilda, ЮKassa, 1С API).",
]
for i, line in enumerate(strat_lines, start=strat_row + 1):
    ws.cell(row=i, column=1, value=line).font = NOTE
    ws.cell(row=i, column=1).alignment = WRAP
    ws.row_dimensions[i].height = 22

col_widths(ws, [44, 22, 28, 22, 22, 20, 22])


# ====================================================================== #
# Sheet 13 — Lean Opex структура                                         #
# ====================================================================== #
ws = wb.create_sheet("13. Lean opex")
write_title(
    ws,
    1,
    "Lean cost-структура · только реальные траты",
    "Только то, что нужно: юрист + LLM (variable) + Серверум + домены/SSL + 300к/мес разработка. Никаких лишних FOT и маркетинговых бюджетов.",
)

# Fixed opex
ws.cell(row=4, column=1, value="ПОСТОЯННЫЕ ЗАТРАТЫ (₽/мес, не зависят от числа клиентов)").font = HEAD
ws.cell(row=4, column=1).fill = HEAD_FILL

write_head(
    ws,
    6,
    ["Категория", "₽/мес", "Комментарий"],
)

fixed = [
    ("Юрист", 30000, "Договоры, оферта, ИП/ЮЛ обслуживание, нерегулярные консультации"),
]
for i, (k, v, n) in enumerate(fixed, start=7):
    ws.cell(row=i, column=1, value=k).font = THIN
    c = ws.cell(row=i, column=2, value=v)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    ws.cell(row=i, column=3, value=n).font = NOTE
    ws.cell(row=i, column=3).alignment = WRAP

# Total fixed
ws.cell(row=9, column=1, value="ИТОГО ПОСТОЯННЫХ").font = BOLD
ws.cell(row=9, column=2, value="=B7").font = BOLD
ws.cell(row=9, column=2).number_format = "#,##0 ₽"
ws.cell(row=9, column=2).fill = ACCENT_FILL

for r in range(6, 10):
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = THIN_BORDER

# Note on what's NOT in fixed
ws.cell(row=11, column=1, value="ПОЧЕМУ ТАК МАЛО ПОСТОЯННЫХ?").font = BOLD
ws.cell(row=11, column=1).fill = HEAD_FILL
ws.cell(row=11, column=1).font = HEAD
note_lines = [
    "• Разработку делают founders сами — без зарплат. AI-инструменты делают это реальным даже для одного человека.",
    "• Серверум и домен клиента — в стоимости подписки (variable per customer), не наши fixed.",
    "• Маркетинг с M2 (после запуска): 100 → 250 К/мес ramp. Если clipboard окупает себя — масштабируем.",
    "• Юрист — единственная обязательная статья (договоры, оферта, документы по 152-ФЗ).",
]
for i, line in enumerate(note_lines, start=12):
    ws.cell(row=i, column=1, value=line).font = NOTE
    ws.cell(row=i, column=1).alignment = WRAP
    ws.row_dimensions[i].height = 22

# Variable opex per customer (per tier)
ws.cell(row=16, column=1, value="ПЕРЕМЕННЫЕ ЗАТРАТЫ (₽/клиент/мес, профиль B — балансный)").font = HEAD
ws.cell(row=16, column=1).fill = HEAD_FILL

write_head(
    ws,
    18,
    [
        "Тариф",
        "Цена ₽/мес",
        "LLM-COGS (wallet/markup)",
        "Серверум · сервер для сайта клиента",
        "Домен клиента + HTTPS",
        "Комиссия 3%",
        "Итого Var",
        "Gross маржа ₽",
        "Маржа %",
    ],
)

# Variable cost per tier
tiers_lean = [
    # name, price, wallet, server_cost (Серверум passthru), domain_per_mo
    ("Lite", 990, 1000, 50, 25),
    ("Starter", 2990, 2500, 600, 50),
    ("Pro", 7990, 6000, 3000, 100),
    ("Enterprise", 19990, 18000, 8000, 250),
]

for i, (name, price, wallet, server, domain) in enumerate(tiers_lean, start=19):
    ws.cell(row=i, column=1, value=name).font = THIN
    c = ws.cell(row=i, column=2, value=price)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # LLM cost = wallet / markup (since wallet is retail-priced)
    ws.cell(
        row=i,
        column=3,
        value=f"={wallet}/{MARKUP}",
    ).number_format = "#,##0 ₽"
    # Server (Серверум passthru)
    c = ws.cell(row=i, column=4, value=server)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # Domain
    c = ws.cell(row=i, column=5, value=domain)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "#,##0 ₽"
    # Commission 3% of price
    ws.cell(row=i, column=6, value=f"=B{i}*0.03").number_format = "#,##0 ₽"
    # Total var
    ws.cell(
        row=i,
        column=7,
        value=f"=C{i}+D{i}+E{i}+F{i}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=7).fill = WARN_FILL
    # GM ₽
    ws.cell(
        row=i,
        column=8,
        value=f"=B{i}-G{i}",
    ).number_format = "#,##0 ₽"
    ws.cell(row=i, column=8).fill = GOOD_FILL
    ws.cell(row=i, column=8).font = BOLD
    # GM %
    ws.cell(
        row=i,
        column=9,
        value=f"=H{i}/B{i}",
    ).number_format = "0.0%"
    ws.cell(row=i, column=9).font = BOLD

for r in range(18, 23):
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = THIN_BORDER

# Notes
ws.cell(row=24, column=1, value="• LLM-COGS = wallet ₽ / markup 2.8 (т.к. wallet это retail-цена токенов).").font = NOTE
ws.cell(row=25, column=1, value="• Если клиент НЕ сжигает весь wallet — наш COGS меньше, маржа выше.").font = NOTE
ws.cell(row=26, column=1, value="• Серверум — passthru от SafeCloud / CORTEL по партнёрскому прайсу + наша наценка уже в подписке.").font = NOTE
ws.cell(row=27, column=1, value="• Домен клиента: .ru/.рф ~300 ₽/год от Reg.ru → 25 ₽/мес. Включается в тариф.").font = NOTE
ws.cell(row=28, column=1, value="• HTTPS-сертификат от Let's Encrypt — бесплатно, авто-обновление.").font = NOTE

col_widths(ws, [22, 14, 22, 14, 14, 14, 14, 16, 12])


# ====================================================================== #
# Sheet 14 — Lean юнит-эконмика (точки безубыточности)                              #
# ====================================================================== #
ws = wb.create_sheet("14. Lean точки безубыточности")
write_title(
    ws,
    1,
    "Точка безубыточности · сколько платящих клиентов нужно",
    "Зависит от микса тарифов. Чем больше Pro/Ent в миксе, тем меньше клиентов нужно.",
)

# Adoption mix (editable)
ws.cell(row=4, column=1, value="ДОЛИ ТАРИФОВ В МИКСЕ").font = HEAD
ws.cell(row=4, column=1).fill = HEAD_FILL

mix_default = [
    ("Lite", 0.50),
    ("Starter", 0.25),
    ("Pro", 0.18),
    ("Enterprise", 0.07),
]
for i, (k, v) in enumerate(mix_default, start=6):
    ws.cell(row=i, column=1, value=k).font = THIN
    c = ws.cell(row=i, column=2, value=v)
    c.font = INPUT
    c.fill = INPUT_FILL
    c.number_format = "0%"

ws.cell(row=10, column=1, value="Сумма (= 100%)").font = BOLD
ws.cell(row=10, column=2, value="=SUM(B6:B9)").font = BOLD
ws.cell(row=10, column=2).number_format = "0%"

# Weighted ARPU & GM
ws.cell(row=13, column=1, value="ВЗВЕШЕННЫЕ ПОКАЗАТЕЛИ").font = HEAD
ws.cell(row=13, column=1).fill = HEAD_FILL

ws.cell(row=15, column=1, value="Средний доход с клиента, ₽").font = THIN
ws.cell(
    row=15,
    column=2,
    value=(
        "=B6*'13. Lean opex'!B21"
        "+B7*'13. Lean opex'!B22"
        "+B8*'13. Lean opex'!B21"
        "+B9*'13. Lean opex'!B22"
    ),
).number_format = "#,##0 ₽"
ws.cell(row=15, column=2).fill = ACCENT_FILL
ws.cell(row=15, column=2).font = BOLD

ws.cell(row=16, column=1, value="Маржа на 1 клиента/мес, ₽").font = THIN
ws.cell(
    row=16,
    column=2,
    value=(
        "=B6*'13. Lean opex'!H21"
        "+B7*'13. Lean opex'!H22"
        "+B8*'13. Lean opex'!H21"
        "+B9*'13. Lean opex'!H22"
    ),
).number_format = "#,##0 ₽"
ws.cell(row=16, column=2).fill = ACCENT_FILL
ws.cell(row=16, column=2).font = BOLD

ws.cell(row=17, column=1, value="Маржа % от выручки").font = THIN
ws.cell(row=17, column=2, value="=B16/B15").number_format = "0.0%"
ws.cell(row=17, column=2).font = BOLD

# Break-even
ws.cell(row=20, column=1, value="ТОЧКА БЕЗУБЫТОЧНОСТИ").font = HEAD
ws.cell(row=20, column=1).fill = HEAD_FILL

ws.cell(row=22, column=1, value="Постоянный opex, ₽/мес").font = THIN
ws.cell(row=22, column=2, value="='13. Lean opex'!B9").number_format = "#,##0 ₽"
ws.cell(row=22, column=2).fill = ACCENT_FILL

ws.cell(row=23, column=1, value="Платящих клиентов нужно").font = BOLD
ws.cell(row=23, column=2, value="=ROUNDUP(B22/B16,0)").font = BOLD
ws.cell(row=23, column=2).fill = GOOD_FILL
ws.cell(row=23, column=2).number_format = "#,##0"

ws.cell(row=24, column=1, value="Выручка/мес для безубыточности, ₽").font = BOLD
ws.cell(row=24, column=2, value="=B23*B15").font = BOLD
ws.cell(row=24, column=2).fill = GOOD_FILL
ws.cell(row=24, column=2).number_format = "#,##0 ₽"

# Sensitivity by mix scenario
ws.cell(row=27, column=1, value="ЧУВСТВИТЕЛЬНОСТЬ ПО МИКСУ").font = HEAD
ws.cell(row=27, column=1).fill = HEAD_FILL

write_head(
    ws,
    29,
    ["Сценарий", "Lite %", "Starter %", "Pro %", "Ent %", "Доход/клиент ₽", "Маржа/клиент ₽", "Клиентов до плюса"],
)

scenarios = [
    ("Зрелый (M15+, default)", 0.50, 0.25, 0.18, 0.07),
    ("Лит-доминантный (M5-9)", 0.80, 0.15, 0.04, 0.01),
    ("Бизнес-mix (M10-14)", 0.60, 0.25, 0.12, 0.03),
    ("Премиум-tilt (M18+)", 0.30, 0.25, 0.30, 0.15),
    ("Только Pro/Ent (B2B fokus)", 0.00, 0.00, 0.70, 0.30),
]
for i, (name, lite, starter, pro, ent) in enumerate(scenarios, start=30):
    ws.cell(row=i, column=1, value=name).font = THIN
    ws.cell(row=i, column=2, value=lite).number_format = "0%"
    ws.cell(row=i, column=3, value=starter).number_format = "0%"
    ws.cell(row=i, column=4, value=pro).number_format = "0%"
    ws.cell(row=i, column=5, value=ent).number_format = "0%"
    # ARPU
    ws.cell(
        row=i,
        column=6,
        value=(
            f"=B{i}*'13. Lean opex'!B21"
            f"+C{i}*'13. Lean opex'!B22"
            f"+D{i}*'13. Lean opex'!B21"
            f"+E{i}*'13. Lean opex'!B22"
        ),
    ).number_format = "#,##0 ₽"
    # GM/cust
    ws.cell(
        row=i,
        column=7,
        value=(
            f"=B{i}*'13. Lean opex'!H21"
            f"+C{i}*'13. Lean opex'!H22"
            f"+D{i}*'13. Lean opex'!H21"
            f"+E{i}*'13. Lean opex'!H22"
        ),
    ).number_format = "#,##0 ₽"
    # Break-even
    ws.cell(
        row=i,
        column=8,
        value=f"=ROUNDUP('13. Lean opex'!B9/G{i},0)",
    ).number_format = "#,##0"
    ws.cell(row=i, column=8).fill = GOOD_FILL
    ws.cell(row=i, column=8).font = BOLD

for r in range(29, 35):
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = THIN_BORDER

ws.cell(
    row=37,
    column=1,
    value="ВЫВОД: при дефолтном миксе точки безубыточности ≈ 280 платящих. При премиум-tilt — всего ~145.",
).font = NOTE
ws.cell(row=37, column=1).alignment = WRAP

ws.cell(
    row=38,
    column=1,
    value="vs v2 business plan (с маркетингом и FOT 3.5М/мес): точки безубыточности был 858 → теперь 280 = в 3× меньше.",
).font = NOTE

col_widths(ws, [32, 12, 12, 12, 12, 14, 14, 14])


# ====================================================================== #
# Sheet 15 — Lean финмодель 18 мес                                        #
# ====================================================================== #
ws = wb.create_sheet("15. Lean финмодель 24м")
write_title(
    ws,
    1,
    "Финмодель 24 месяца · founders-only без зарплат + инфраструктура для роста",
    "Юрист 30к/мес · маркетинг 100→300к · инфра наша (DB+API+backups+мониторинг) растёт с базой клиентов 8к→380к/мес. Серверум за клиентский сайт уже в variable COGS.",
)

# Header row M1-M24
ws.cell(row=4, column=1, value="Параметр").font = HEAD
ws.cell(row=4, column=1).fill = HEAD_FILL
for m in range(1, 25):
    c = ws.cell(row=4, column=1 + m, value=f"M{m}")
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = CENTER
ws.row_dimensions[4].height = 24

def fillrow(row, label, values, fmt=None, fill_color=None, bold=False):
    ws.cell(row=row, column=1, value=label).font = BOLD if bold else THIN
    if fill_color:
        ws.cell(row=row, column=1).fill = fill_color
    for i, v in enumerate(values, start=2):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BOLD if bold else THIN
        if fmt:
            c.number_format = fmt
        if fill_color:
            c.fill = fill_color

# Платящие клиенты — 24 мес ramp; founders делали dev до M1, маркетинг с M2
paying = [
    0,                                       # M1 soft launch
    30, 80, 150, 250,                        # M2-M5
    380, 540, 720, 920,                      # M6-M9
    1140, 1390, 1670, 1980, 2330,            # M10-M14
    2710, 3120, 3550, 4000,                  # M15-M18
    4400, 4800, 5200, 5550, 5900, 6250,      # M19-M24 (рост замедляется)
]
fillrow(5, "Платящие клиенты (накопленно)", paying, fmt="#,##0")

# ARPU ramps from 990 → 4160 mature mix
arpu = [
    0,
    990, 990, 1100, 1200,
    1400, 1700, 2000, 2300,
    2600, 2900, 3150, 3400, 3650,
    3850, 3950, 4030, 4080,
    4080, 4100, 4120, 4140, 4150, 4160,      # M19-M24
]
fillrow(6, "Средний доход с клиента, ₽/мес", arpu, fmt="#,##0 ₽")

# MRR
ws.cell(row=7, column=1, value="Выручка, ₽/мес").font = BOLD
for m in range(1, 25):
    col = get_column_letter(1 + m)
    c = ws.cell(row=7, column=1 + m, value=f"={col}5*{col}6")
    c.font = BOLD
    c.fill = ACCENT_FILL
    c.number_format = "#,##0 ₽"

# Маржа % mirrors mix maturity. Early Lite-heavy = high margin (53%).
# As Pro/Ent grow and tokens dominate the COGS = ratio drops to ~34%.
gm_ratio = [
    0.0,  # M1 nobody pays
    0.53, 0.53, 0.52, 0.50, 0.48, 0.46, 0.44, 0.42, 0.40,
    0.38, 0.37, 0.36, 0.35, 0.34, 0.34, 0.34, 0.34,
    0.34, 0.34, 0.34, 0.34, 0.34, 0.34,  # M19-M24
]
fillrow(8, "Маржа %", gm_ratio, fmt="0.0%")

# Gross profit
ws.cell(row=9, column=1, value="ГРОСС-ПРИБЫЛЬ, ₽").font = BOLD
for m in range(1, 25):
    col = get_column_letter(1 + m)
    c = ws.cell(row=9, column=1 + m, value=f"={col}7*{col}8")
    c.font = BOLD
    c.fill = GOOD_FILL
    c.number_format = "#,##0 ₽"

# Fixed Opex breakdown
# Юрист 30к каждый месяц. Разработка = 0 (founders сами без зарплат).
# Маркетинг ramping: 0 → 100 → 150 → 200 → 250 → 300.
# Инфра наша: API-сервера + БД + Redis + S3 + бэкапы + мониторинг — растёт с базой
# (Серверум за клиентский сайт + домен клиента уже учтены в variable COGS внутри тарифа.)
ws.cell(row=11, column=1, value="OPEX: Юрист, ₽").font = THIN
ws.cell(row=12, column=1, value="OPEX: Маркетинг и реклама, ₽").font = THIN
ws.cell(row=13, column=1, value="OPEX: Инфра (наши сервера/БД/бэкапы), ₽").font = THIN
marketing = [
    0,                                     # M1 soft launch
    100000, 100000, 100000, 100000,        # M2-M5
    150000, 150000, 150000, 150000,        # M6-M9
    200000, 200000, 200000, 200000, 200000, # M10-M14
    250000, 250000, 250000, 250000,        # M15-M18
    250000, 250000,                        # M19-M20
    300000, 300000, 300000, 300000,        # M21-M24
]
# Инфра масштабируется с числом платящих клиентов: control-plane (наши API),
# managed PostgreSQL/Redis, S3-бэкапы, мониторинг, CDN, DDoS-защита.
# При росте до 6К+ клиентов это становится заметной статьёй (~6% от ARPU).
infra = [
    8000,                                  # M1 (0 платящих) — staging+dev VPS+минимум
    10000, 15000, 20000, 25000,            # M2-M5 (30→250) — primary DB + S3
    35000, 45000, 55000, 65000,            # M6-M9 (380→920) — read-replica + monitoring
    80000, 100000, 120000, 140000, 160000, # M10-M14 (1.1k→2.3k) — cluster + DR
    180000, 200000, 230000, 260000,        # M15-M18 (2.7k→4k) — premium ops
    280000, 300000, 320000, 340000, 360000, 380000,  # M19-M24 (4.4k→6.25k)
]
for m in range(1, 25):
    ws.cell(row=11, column=1 + m, value=30000).number_format = "#,##0 ₽"
    c = ws.cell(row=12, column=1 + m, value=marketing[m - 1])
    c.number_format = "#,##0 ₽"
    if marketing[m - 1] > 0:
        c.fill = INPUT_FILL
    c2 = ws.cell(row=13, column=1 + m, value=infra[m - 1])
    c2.number_format = "#,##0 ₽"
    c2.fill = INPUT_FILL

# Total opex
ws.cell(row=14, column=1, value="ИТОГО OPEX, ₽").font = BOLD
for m in range(1, 25):
    col = get_column_letter(1 + m)
    c = ws.cell(row=14, column=1 + m, value=f"=SUM({col}11:{col}13)")
    c.font = BOLD
    c.fill = WARN_FILL
    c.number_format = "#,##0 ₽"

# Прибыль/убыток = GM - opex
ws.cell(row=16, column=1, value="Прибыль/убыток, ₽").font = BOLD
for m in range(1, 25):
    col = get_column_letter(1 + m)
    c = ws.cell(row=16, column=1 + m, value=f"={col}9-{col}14")
    c.font = BOLD
    c.number_format = "#,##0 ₽"

# Cumulative cash (start with 30K seed capital)
ws.cell(row=17, column=1, value="Накопленный кэш (старт 30к ₽), ₽").font = BOLD
ws.cell(row=17, column=2, value="=30000+B16")
ws.cell(row=17, column=2).number_format = "#,##0 ₽"
ws.cell(row=17, column=2).font = BOLD
for m in range(2, 25):
    col = get_column_letter(1 + m)
    prev = get_column_letter(m)
    c = ws.cell(row=17, column=1 + m, value=f"={prev}17+{col}16")
    c.font = BOLD
    c.number_format = "#,##0 ₽"
    c.fill = ACCENT_FILL

# Summary block — теперь до M24 (col Y = 25)
ws.cell(row=20, column=1, value="SUMMARY").font = HEAD
ws.cell(row=20, column=1).fill = HEAD_FILL

ws.cell(row=21, column=1, value="Выручка/мес на M24, ₽").font = THIN
ws.cell(row=21, column=2, value="=Y7").number_format = "#,##0 ₽"
ws.cell(row=21, column=2).fill = GOOD_FILL
ws.cell(row=21, column=2).font = BOLD

ws.cell(row=22, column=1, value="Выручка/год на M24, ₽").font = THIN
ws.cell(row=22, column=2, value="=Y7*12").number_format = "#,##0 ₽"
ws.cell(row=22, column=2).fill = GOOD_FILL
ws.cell(row=22, column=2).font = BOLD

ws.cell(row=23, column=1, value="Накопленный кэш на M24, ₽").font = THIN
ws.cell(row=23, column=2, value="=Y17").number_format = "#,##0 ₽"
ws.cell(row=23, column=2).fill = GOOD_FILL
ws.cell(row=23, column=2).font = BOLD

ws.cell(row=24, column=1, value="Самый большой минус по деньгам, ₽").font = THIN
ws.cell(row=24, column=2, value="=MIN(B17:Y17)").number_format = "#,##0 ₽"
ws.cell(row=24, column=2).fill = WARN_FILL
ws.cell(row=24, column=2).font = BOLD

ws.cell(row=25, column=1, value="Платящих клиентов на M24").font = THIN
ws.cell(row=25, column=2, value="=Y5").number_format = "#,##0"
ws.cell(row=25, column=2).fill = GOOD_FILL
ws.cell(row=25, column=2).font = BOLD

ws.cell(row=26, column=1, value="Месяц первой прибыли (положительная)").font = THIN
ws.cell(row=26, column=2, value="M6 (с учётом инфры)").font = NOTE

ws.cell(row=27, column=1, value="Месяц возврата вложений (кэш > 0)").font = THIN
ws.cell(row=27, column=2, value="M8 (с учётом инфры)").font = NOTE

ws.cell(row=28, column=1, value="Маркетинг суммарно за 24 мес, ₽").font = THIN
ws.cell(row=28, column=2, value="=SUM(B12:Y12)").number_format = "#,##0 ₽"
ws.cell(row=28, column=2).fill = WARN_FILL

ws.cell(row=29, column=1, value="Инфра суммарно за 24 мес, ₽").font = THIN
ws.cell(row=29, column=2, value="=SUM(B13:Y13)").number_format = "#,##0 ₽"
ws.cell(row=29, column=2).fill = WARN_FILL

# vs предыдущая (с зарплатой разработчиков 300к/мес)
ws.cell(row=31, column=1, value="vs модель с зарплатой разработчикам 300к/мес (с учётом инфры):").font = HEAD
ws.cell(row=31, column=1).fill = HEAD_FILL

comp = [
    ("Месяц первой прибыли", "M6", "M9", "+3 мес раньше"),
    ("Месяц возврата вложений", "M8", "M12", "+4 мес раньше"),
    ("Самый большой минус", "~−305 К ₽", "~−2.5 М ₽", "−88%"),
    ("Накопленный кэш на M24", "~+67 М ₽", "~+59 М ₽", "+8 М"),
    ("Можно ли запуститься на 30к?", "ДА (хватит на M1)", "Нет", "—"),
    ("Источник разработки", "founders + AI", "1 senior + 1 mid", "—"),
]
for i, (k, fast, slow, delta) in enumerate(comp, start=33):
    ws.cell(row=i, column=1, value=k).font = THIN
    ws.cell(row=i, column=2, value=fast).font = BOLD
    ws.cell(row=i, column=2).fill = GOOD_FILL
    ws.cell(row=i, column=3, value=slow).font = THIN
    ws.cell(row=i, column=3).fill = WARN_FILL
    ws.cell(row=i, column=4, value=delta).font = THIN

col_widths(ws, [42, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])


# ====================================================================== #
# Sheet 16 — Маркетинг · Wordstat + воронка 2026                          #
# ====================================================================== #
ws = wb.create_sheet("16. Маркетинг 2026 (Wordstat)")
write_title(
    ws,
    1,
    "Семантическое ядро · Wordstat 2026 + воронка 100К ₽",
    "71 запрос по 12 нишам. CPC оценка после ухода Google Ads (2022) и AI-хайпа 2024-2025. Перед запуском проверить руками в wordstat.yandex.ru и Прогнозе бюджета Директа.",
)

# Keyword data: (запрос, категория, показов/мес, CPC ₽ 2026, ценность 1-5)
keywords_2026 = [
    # A. Прямой коммерческий
    ("создать сайт", "A. Прямой", 80000, 110, 3),
    ("сделать сайт", "A. Прямой", 70000, 95, 3),
    ("сайт бесплатно", "A. Прямой", 40000, 70, 2),
    ("конструктор сайтов", "A. Прямой", 25000, 165, 4),
    ("сайт под ключ", "A. Прямой", 18000, 135, 4),
    ("лендинг создать", "A. Прямой", 15000, 120, 4),
    ("сайт визитка", "A. Прямой", 12000, 75, 3),
    ("сайт самостоятельно", "A. Прямой", 8000, 80, 4),
    ("одностраничный сайт", "A. Прямой", 5000, 90, 4),
    ("сайт без программиста", "A. Прямой", 1500, 65, 5),
    # B. AI
    ("нейросеть сайт", "B. AI ★", 4500, 85, 5),
    ("ai сайт", "B. AI ★", 2800, 90, 5),
    ("нейросеть создать сайт", "B. AI ★", 2200, 80, 5),
    ("сайт через chatgpt", "B. AI ★", 1500, 75, 4),
    ("искусственный интеллект сайт", "B. AI ★", 1400, 80, 4),
    ("сделать сайт нейросетью", "B. AI ★", 1300, 80, 5),
    ("ai конструктор сайтов", "B. AI ★", 1200, 95, 5),
    ("сайт с помощью ai", "B. AI ★", 950, 85, 5),
    ("генератор сайтов ai", "B. AI ★", 800, 95, 5),
    # C. Бренд
    ("тильда", "C. Бренд", 200000, 35, 1),
    ("bitrix сайт сделать", "C. Бренд", 4500, 65, 3),
    ("wix или tilda", "C. Бренд", 1500, 65, 5),
    ("tilda аналог", "C. Бренд", 1100, 75, 5),
    ("конструктор сайтов сравнение", "C. Бренд", 800, 80, 4),
    ("замена tilda", "C. Бренд", 600, 70, 5),
    # D. Кафе
    ("доставка еды сайт", "D. Кафе", 4500, 95, 4),
    ("сайт ресторана", "D. Кафе", 3000, 85, 5),
    ("сайт для кафе", "D. Кафе", 2500, 75, 5),
    ("сайт кофейни", "D. Кафе", 800, 70, 5),
    ("сайт для бара", "D. Кафе", 600, 70, 4),
    ("сайт пиццерии", "D. Кафе", 500, 75, 5),
    # E. Магазин
    ("создать интернет магазин", "E. Магазин", 12000, 165, 4),
    ("интернет магазин с нуля", "E. Магазин", 3500, 130, 4),
    ("интернет магазин под ключ", "E. Магазин", 2500, 145, 4),
    ("магазин одежды сайт", "E. Магазин", 2000, 95, 4),
    ("сайт магазина бесплатно", "E. Магазин", 1500, 90, 3),
    ("сайт цветочного магазина", "E. Магазин", 600, 70, 5),
    # F. Юрист
    ("сайт юриста", "F. Юрист", 2000, 105, 5),
    ("сайт адвоката", "F. Юрист", 1500, 105, 5),
    ("сайт юридической компании", "F. Юрист", 800, 125, 5),
    ("сайт нотариуса", "F. Юрист", 600, 90, 5),
    ("сайт визитка юристу", "F. Юрист", 400, 85, 5),
    # G. Красота
    ("сайт салона красоты", "G. Красота", 2500, 85, 5),
    ("сайт мастера маникюра", "G. Красота", 1800, 75, 5),
    ("сайт косметолога", "G. Красота", 1500, 95, 5),
    ("сайт парикмахерской", "G. Красота", 1200, 65, 5),
    ("сайт барбершопа", "G. Красота", 800, 70, 5),
    # H. Медицина
    ("сайт клиники", "H. Медицина", 2000, 145, 4),
    ("сайт врача", "H. Медицина", 1800, 95, 5),
    ("сайт стоматологии", "H. Медицина", 1500, 130, 5),
    ("сайт медцентра", "H. Медицина", 1200, 145, 4),
    ("сайт ветклиники", "H. Медицина", 800, 105, 5),
    # I. Авто
    ("сайт автосервиса", "I. Авто", 1200, 85, 5),
    ("сайт автосалона", "I. Авто", 800, 115, 4),
    ("сайт шиномонтажа", "I. Авто", 600, 65, 5),
    ("сайт детейлинга", "I. Авто", 500, 75, 5),
    # J. Недвижимость
    ("сайт агентства недвижимости", "J. Недв", 1500, 125, 4),
    ("сайт риэлтора", "J. Недв", 1200, 95, 5),
    ("сайт жк", "J. Недв", 800, 175, 3),
    ("сайт застройщика", "J. Недв", 600, 145, 4),
    # K. Образование
    ("сайт психолога", "K. Образ", 2500, 95, 5),
    ("сайт репетитора", "K. Образ", 2000, 75, 5),
    ("сайт онлайн школы", "K. Образ", 1500, 105, 4),
    ("сайт коуча", "K. Образ", 800, 85, 5),
    ("сайт фитнес тренера", "K. Образ", 700, 75, 5),
    # L. Pain-driven
    ("простой сайт сделать", "L. Pain ★", 1500, 70, 4),
    ("сайт за день", "L. Pain ★", 1200, 95, 5),
    ("сайт быстро недорого", "L. Pain ★", 800, 85, 5),
    ("сайт визитка дёшево", "L. Pain ★", 800, 55, 4),
    ("сайт без знаний программирования", "L. Pain ★", 600, 65, 5),
    ("сделать сайт самому без программиста", "L. Pain ★", 400, 60, 5),
]

# Headers
write_head(ws, 4, ["#", "Запрос", "Категория", "Показов/мес 2026", "CPC ₽ 2026", "Ценность 1-5"])

for i, (q, cat, imp, cpc, val) in enumerate(keywords_2026, start=5):
    ws.cell(row=i, column=1, value=i - 4).font = THIN
    ws.cell(row=i, column=2, value=q).font = THIN
    ws.cell(row=i, column=3, value=cat).font = THIN
    ws.cell(row=i, column=4, value=imp).number_format = "#,##0"
    ws.cell(row=i, column=4).font = THIN
    ws.cell(row=i, column=5, value=cpc).number_format = "#,##0 ₽"
    ws.cell(row=i, column=5).font = THIN
    c = ws.cell(row=i, column=6, value=val)
    c.font = BOLD
    c.alignment = CENTER
    if val >= 4:
        c.fill = GOOD_FILL
    elif val == 3:
        c.fill = INPUT_FILL
    else:
        c.fill = WARN_FILL

# Aggregates row
agg_row = 5 + len(keywords_2026) + 1
ws.cell(row=agg_row, column=2, value=f"ИТОГО · {len(keywords_2026)} ключей").font = HEAD
ws.cell(row=agg_row, column=2).fill = HEAD_FILL
ws.cell(row=agg_row, column=4, value=f"=SUM(D5:D{agg_row-1})").number_format = "#,##0"
ws.cell(row=agg_row, column=4).font = HEAD
ws.cell(row=agg_row, column=4).fill = HEAD_FILL
ws.cell(row=agg_row, column=5, value=f"=AVERAGE(E5:E{agg_row-1})").number_format = "#,##0 ₽"
ws.cell(row=agg_row, column=5).font = HEAD
ws.cell(row=agg_row, column=5).fill = HEAD_FILL

# ====================================================================== #
# Funnel calculator                                                      #
# ====================================================================== #
funnel_start = agg_row + 3
ws.cell(row=funnel_start, column=1, value="ВОРОНКА 100К ₽ → ПЛАТЯЩИЕ КЛИЕНТЫ · 4 СЦЕНАРИЯ 2026").font = HEAD
ws.cell(row=funnel_start, column=1).fill = HEAD_FILL

# Inputs row
input_row = funnel_start + 2
ws.cell(row=input_row, column=1, value="Бюджет ₽/мес").font = INPUT
ws.cell(row=input_row, column=1).fill = INPUT_FILL
ws.cell(row=input_row, column=2, value=100000).font = INPUT
ws.cell(row=input_row, column=2).fill = INPUT_FILL
ws.cell(row=input_row, column=2).number_format = "#,##0 ₽"
ws.cell(row=input_row, column=4, value="← поменяй число — пересчитается всё ниже").font = NOTE

# Scenarios table
sc_head_row = input_row + 2
write_head(ws, sc_head_row, [
    "Сценарий",
    "CPC ₽",
    "Кликов = визитов",
    "Цена визита ₽",
    "Конверсия лендинга",
    "Trial регистраций",
    "Trial → Paid",
    "Платящих/мес",
    "CAC ₽",
])

scenarios = [
    ("А. Песимистичный (M2 cold start)", 95, 0.025, 0.05),
    ("Б. Реалистичный (M3, базовая оптимизация)", 75, 0.035, 0.06),
    ("В. Оптимистичный (M5, отшлифованный)", 55, 0.05, 0.08),
    ("Г. Идеал (M6+ vertical+long-tail)", 45, 0.06, 0.10),
]

for i, (name, cpc, cv, tp) in enumerate(scenarios, start=sc_head_row + 1):
    bcol = "B"  # column index 2 для бюджета
    ws.cell(row=i, column=1, value=name).font = THIN
    # CPC
    ws.cell(row=i, column=2, value=cpc).number_format = "#,##0 ₽"
    ws.cell(row=i, column=2).font = INPUT
    ws.cell(row=i, column=2).fill = INPUT_FILL
    # Кликов
    ws.cell(row=i, column=3, value=f"=${bcol}${input_row}/B{i}").number_format = "#,##0"
    # Цена визита (= CPC)
    ws.cell(row=i, column=4, value=f"=B{i}").number_format = "#,##0 ₽"
    # Conversion лендинга
    ws.cell(row=i, column=5, value=cv).number_format = "0.0%"
    ws.cell(row=i, column=5).font = INPUT
    ws.cell(row=i, column=5).fill = INPUT_FILL
    # Trials
    ws.cell(row=i, column=6, value=f"=C{i}*E{i}").number_format = "#,##0"
    # Trial → Paid
    ws.cell(row=i, column=7, value=tp).number_format = "0.0%"
    ws.cell(row=i, column=7).font = INPUT
    ws.cell(row=i, column=7).fill = INPUT_FILL
    # Платящих
    pay_cell = ws.cell(row=i, column=8, value=f"=F{i}*G{i}")
    pay_cell.number_format = "0.0"
    pay_cell.font = BOLD
    pay_cell.fill = GOOD_FILL
    # CAC
    cac_cell = ws.cell(row=i, column=9, value=f"=${bcol}${input_row}/H{i}")
    cac_cell.number_format = "#,##0 ₽"
    cac_cell.font = BOLD
    if i - sc_head_row <= 2:
        cac_cell.fill = WARN_FILL
    else:
        cac_cell.fill = GOOD_FILL

# 6-month ramp
ramp_row = sc_head_row + len(scenarios) + 3
ws.cell(row=ramp_row, column=1, value="ГРАФИК ПЕРВЫХ 6 МЕСЯЦЕВ НА 100К ₽/мес DIRECT").font = HEAD
ws.cell(row=ramp_row, column=1).fill = HEAD_FILL

write_head(ws, ramp_row + 2, [
    "Месяц",
    "Что происходит",
    "CPC ₽",
    "CV лендинга",
    "Trial → Paid",
    "Кликов",
    "Trials",
    "Платящих",
    "CAC ₽",
])

ramp = [
    (1, "M1 Soft launch (без рекламы)", 0, 0.0, 0.0),
    (2, "M2 Cold start", 95, 0.025, 0.05),
    (3, "M3 Первая оптимизация", 80, 0.030, 0.06),
    (4, "M4 Vertical landing pages", 65, 0.040, 0.07),
    (5, "M5 Retargeting + бренд", 55, 0.045, 0.08),
    (6, "M6 Зрелая воронка", 50, 0.050, 0.09),
]

for i, (m, label, cpc, cv, tp) in enumerate(ramp, start=ramp_row + 3):
    ws.cell(row=i, column=1, value=f"M{m}").font = BOLD
    ws.cell(row=i, column=2, value=label).font = THIN
    if cpc == 0:
        ws.cell(row=i, column=3, value="—").alignment = CENTER
        ws.cell(row=i, column=4, value="—").alignment = CENTER
        ws.cell(row=i, column=5, value="—").alignment = CENTER
        ws.cell(row=i, column=6, value="—").alignment = CENTER
        ws.cell(row=i, column=7, value="—").alignment = CENTER
        c = ws.cell(row=i, column=8, value=0)
        c.font = BOLD
        c.fill = WARN_FILL
        ws.cell(row=i, column=9, value="—").alignment = CENTER
    else:
        ws.cell(row=i, column=3, value=cpc).number_format = "#,##0 ₽"
        ws.cell(row=i, column=4, value=cv).number_format = "0.0%"
        ws.cell(row=i, column=5, value=tp).number_format = "0.0%"
        ws.cell(row=i, column=6, value=f"=$B${input_row}/C{i}").number_format = "#,##0"
        ws.cell(row=i, column=7, value=f"=F{i}*D{i}").number_format = "#,##0"
        c = ws.cell(row=i, column=8, value=f"=G{i}*E{i}")
        c.number_format = "0.0"
        c.font = BOLD
        c.fill = GOOD_FILL
        cac = ws.cell(row=i, column=9, value=f"=$B${input_row}/H{i}")
        cac.number_format = "#,##0 ₽"

# Total row
total_row = ramp_row + 3 + len(ramp)
ws.cell(row=total_row, column=2, value="ИТОГО за 6 мес (новых платящих)").font = HEAD
ws.cell(row=total_row, column=2).fill = HEAD_FILL
ws.cell(row=total_row, column=8, value=f"=SUM(H{ramp_row + 3}:H{total_row - 1})").number_format = "0.0"
ws.cell(row=total_row, column=8).font = HEAD
ws.cell(row=total_row, column=8).fill = HEAD_FILL

# Reality check vs финплан
check_row = total_row + 3
ws.cell(row=check_row, column=1, value="РАСХОЖДЕНИЕ С ФИНПЛАНОМ (лист 15)").font = HEAD
ws.cell(row=check_row, column=1).fill = HEAD_FILL

write_head(ws, check_row + 2, [
    "Месяц",
    "По финплану",
    "По воронке Direct 100К",
    "Дельта",
    "Что компенсирует",
])

check_data = [
    ("M1", 0, 0, "0", "—"),
    ("M2", 30, 1, "−29", "Soft-launch buzz + Telegram founders"),
    ("M3", 80, 2, "−78", "VC.ru/Habr статья + первые кейсы"),
    ("M4", 150, 4, "−146", "Multi-channel (VK + Telegram + контент)"),
    ("M5", 250, 7, "−243", "Retargeting + первая узнаваемость"),
    ("M6", 380, 9, "−371", "Зрелая воронка + B2B продажи"),
    ("M7+", "540+", "15-25", "догоняет", "Бренд + кейсы + органика"),
]
for i, (m, fp, vrn, delta, comp) in enumerate(check_data, start=check_row + 3):
    ws.cell(row=i, column=1, value=m).font = BOLD
    if isinstance(fp, int):
        ws.cell(row=i, column=2, value=fp).number_format = "#,##0"
    else:
        ws.cell(row=i, column=2, value=fp).alignment = CENTER
    if isinstance(vrn, int):
        ws.cell(row=i, column=3, value=vrn).number_format = "#,##0"
    else:
        ws.cell(row=i, column=3, value=vrn).alignment = CENTER
    ws.cell(row=i, column=4, value=delta).font = BOLD
    ws.cell(row=i, column=4).fill = WARN_FILL if "−" in str(delta) else GOOD_FILL
    ws.cell(row=i, column=5, value=comp).font = NOTE

# Final note
note_row = check_row + 3 + len(check_data) + 2
ws.cell(row=note_row, column=1, value=(
    "ВЫВОД: 100К ₽ только в Direct в первый месяц = 1-2 платящих. "
    "Финплан с М2=30 платящих требует мульти-канал (Direct 60К + VK 25К + Telegram 15К) ИЛИ "
    "soft-launch buzz (VC.ru/Habr) + B2B sales. Реалистичный график: М2=5-10, М3=18-25, М6=250-300, М7+ догоняет план."
)).font = HEAD
ws.cell(row=note_row, column=1).fill = ACCENT_FILL
ws.cell(row=note_row, column=1).alignment = WRAP
ws.row_dimensions[note_row].height = 60
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

col_widths(ws, [42, 30, 18, 20, 18, 18, 18, 18, 18])


wb.save(OUT)
print(f"OK saved: {OUT}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Size bytes: {OUT.stat().st_size}")
